# -*- coding: utf-8 -*-
"""导出高德云图客流分析+常驻人口画像数据到Excel"""
import json, os, sys
sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
except ImportError:
    os.system(f'{sys.executable} -m pip install openpyxl')
    import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
hdr_font = Font(bold=True, color="FFFFFF", size=11, name='Microsoft YaHei')
hdr_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
body_font = Font(size=10, name='Microsoft YaHei')
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
title_font = Font(bold=True, size=13, name='Microsoft YaHei')
sub_font = Font(bold=True, size=11, name='Microsoft YaHei', color='1A2744')

# ═══ 数据 ═══
# 客流分析 - 全部日
flow_allday = {
    "uv": 2284077,
    "age": {"19-24": 0.055584, "25-29": 0.103054, "30-34": 0.123390, "35-39": 0.171978,
            "40-44": 0.160421, "45-49": 0.133867, "50-54": 0.116163, "55-59": 0.075949,
            "60-64": 0.038170, "65+": 0.021425},
    "sex": {"男": 0.557181, "女": 0.442819},
    "consumption_level": {"低": 0.115552, "中": 0.456174, "中高": 0.299882, "高": 0.128392},
    "dining_consum_level": {"0-49": 0.587863, "50-99": 0.356484, "100-199": 0.047463,
                            "200-299": 0.005245, "300-399": 0.001127, "400-499": 0.000357, "500+": 0.001460},
    "hotel_consume_level": {"0-99": 0.417213, "100-299": 0.422564, "300-499": 0.038623,
                            "500-999": 0.043398, "1000+": 0.078202},
    "travel_tool": {"火车": 0.659788, "飞机": 0.244461, "汽车": 0.095751},
    "trip_mode_r3m": {"驾车": 0.445922, "火车": 0.157540, "打车": 0.105995, "地铁": 0.084231,
                      "公交": 0.070683, "骑行": 0.052653, "步行": 0.049446, "飞机": 0.029184,
                      "摩托": 0.003496, "长途汽车": 0.000850},
}

flow_holiday = {
    "uv": 2316671,
    "age": {"19-24": 0.057758, "25-29": 0.106783, "30-34": 0.124313, "35-39": 0.172007,
            "40-44": 0.159243, "45-49": 0.132258, "50-54": 0.114411, "55-59": 0.074762,
            "60-64": 0.037489, "65+": 0.020976},
    "sex": {"男": 0.552313, "女": 0.447687},
    "consumption_level": {"低": 0.115591, "中": 0.455172, "中高": 0.298572, "高": 0.130665},
    "hotel_consume_level": {"0-99": 0.416239, "100-299": 0.421370, "300-499": 0.038712,
                            "500-999": 0.044225, "1000+": 0.079455},
    "travel_tool": {"火车": 0.657483, "飞机": 0.245381, "汽车": 0.097136},
}

flow_weekend = {
    "uv": 2313960,
    "age": {"19-24": 0.057377, "25-29": 0.106381, "30-34": 0.124172, "35-39": 0.171923,
            "40-44": 0.159296, "45-49": 0.132519, "50-54": 0.114728, "55-59": 0.074987,
            "60-64": 0.037607, "65+": 0.021010},
    "sex": {"男": 0.552800, "女": 0.447200},
    "consumption_level": {"低": 0.115206, "中": 0.455093, "中高": 0.298952, "高": 0.130749},
    "hotel_consume_level": {"0-99": 0.415851, "100-299": 0.421668, "300-499": 0.038741,
                            "500-999": 0.044194, "1000+": 0.079546},
    "travel_tool": {"火车": 0.657585, "飞机": 0.245364, "汽车": 0.097051},
}

# 常驻分析
pmnt_active = {
    "uv": 3530838,
    "age": {"19-24": 0.058667, "25-29": 0.110651, "30-34": 0.128968, "35-39": 0.174976,
            "40-44": 0.156799, "45-49": 0.126881, "50-54": 0.108120, "55-59": 0.071593,
            "60-64": 0.038641, "65+": 0.024705},
    "sex": {"男": 0.555310, "女": 0.444690},
    "consumption_level": {"低": 0.121373, "中": 0.477497, "中高": 0.273113, "高": 0.128017},
    "occupation": {"公司职员": 0.490137, "商业服务职员": 0.292720, "其他": 0.122357,
                   "医护人员": 0.049821, "生产运输职员": 0.035127, "教职工": 0.006533,
                   "公司高管": 0.002938, "自由职业": 0.000367},
    "education": {"本科": 0.004154, "专科": 0.000879, "硕士": 0.000774, "博士": 0.000086},
}

pmnt_home = {
    "uv": 1679214,
    "age": {"19-24": 0.058436, "25-29": 0.107919, "30-34": 0.125151, "35-39": 0.173410,
            "40-44": 0.156833, "45-49": 0.127430, "50-54": 0.110144, "55-59": 0.074588,
            "60-64": 0.040539, "65+": 0.025550},
    "sex": {"男": 0.530056, "女": 0.469944},
    "consumption_level": {"低": 0.122832, "中": 0.487279, "中高": 0.272990, "高": 0.116899},
    "occupation": {"公司职员": 0.484439, "商业服务职员": 0.299750, "其他": 0.130107,
                   "医护人员": 0.051374, "生产运输职员": 0.025456, "教职工": 0.006783,
                   "公司高管": 0.001700, "自由职业": 0.000390},
    "education": {"本科": 0.004773, "专科": 0.001006, "硕士": 0.000879, "博士": 0.000104},
}

pmnt_work = {
    "uv": 1044283,
    "age": {"19-24": 0.056121, "25-29": 0.114026, "30-34": 0.133805, "35-39": 0.176445,
            "40-44": 0.159752, "45-49": 0.129473, "50-54": 0.108924, "55-59": 0.069298,
            "60-64": 0.033420, "65+": 0.018736},
    "sex": {"男": 0.517941, "女": 0.482059},
    "consumption_level": {"低": 0.092429, "中": 0.450464, "中高": 0.328876, "高": 0.128231},
    "occupation": {"公司职员": 0.538183, "商业服务职员": 0.262205, "其他": 0.119906,
                   "医护人员": 0.040822, "生产运输职员": 0.030765, "教职工": 0.005391,
                   "公司高管": 0.002396, "自由职业": 0.000331},
    "education": {"本科": 0.004364, "专科": 0.000885, "硕士": 0.000780, "博士": 0.000084},
}

def write_section(ws, start_row, title, data_dict, categories_order=None):
    """写入一个分析板块"""
    row = start_row
    ws.cell(row=row, column=1, value=title).font = sub_font
    row += 1

    if categories_order is None:
        categories_order = list(data_dict.keys())

    # 表头
    headers = ["类别"] + categories_order
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin_border
    row += 1

    # 获取所有标签
    all_labels = set()
    for cat in categories_order:
        if cat in data_dict and isinstance(data_dict[cat], dict):
            all_labels.update(data_dict[cat].keys())

    # 排序标签
    label_order = sorted(all_labels)

    for label in label_order:
        ws.cell(row=row, column=1, value=label).font = body_font
        ws.cell(row=row, column=1).border = thin_border
        for col, cat in enumerate(categories_order, 2):
            if cat in data_dict and isinstance(data_dict[cat], dict):
                val = data_dict[cat].get(label, 0)
                cell = ws.cell(row=row, column=col, value=f"{val*100:.2f}%")
                cell.font = body_font
                cell.alignment = Alignment(horizontal='center')
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
        row += 1

    return row + 1

# ═══ Sheet 1: 客流分析 ═══
ws1 = wb.active
ws1.title = "客流分析"
ws1.cell(row=1, column=1, value="香洲区(440402)客流画像分析 - 2026年6月").font = title_font
ws1.cell(row=2, column=1, value="数据来源: 高德云图 cloudmap-gateway | 行政区: 珠海市香洲区(含桂山镇)").font = Font(size=9, name='Microsoft YaHei', color='666666')

# UV汇总
row = 4
ws1.cell(row=row, column=1, value="月均天级客流指数(UV)").font = sub_font
row += 1
for col, h in enumerate(["类型", "UV指数", "说明"], 1):
    cell = ws1.cell(row=row, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = thin_border
row += 1
uv_data = [
    ("全部日(月均)", flow_allday["uv"], "整月所有天数的平均每日客流"),
    ("节假日(日均)", flow_holiday["uv"], "仅法定节假日平均每日客流"),
    ("周末(日均)", flow_weekend["uv"], "仅周六日平均每日客流"),
]
for typ, uv, desc in uv_data:
    for col, val in enumerate([typ, f"{uv:,}", desc], 1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.font = body_font
        cell.border = thin_border
    row += 1

row += 1

# 性别
age_data = {"全部日": flow_allday["age"], "节假日": flow_holiday["age"], "周末": flow_weekend["age"]}
row = write_section(ws1, row, "年龄分布", age_data, ["全部日", "节假日", "周末"])

sex_data = {"全部日": flow_allday["sex"], "节假日": flow_holiday["sex"], "周末": flow_weekend["sex"]}
row = write_section(ws1, row, "性别分布", sex_data, ["全部日", "节假日", "周末"])

cons_data = {"全部日": flow_allday["consumption_level"], "节假日": flow_holiday["consumption_level"], "周末": flow_weekend["consumption_level"]}
row = write_section(ws1, row, "消费水平", cons_data, ["全部日", "节假日", "周末"])

dining = {"全部日": flow_allday["dining_consum_level"]}
row = write_section(ws1, row, "餐饮消费水平(元)", dining, ["全部日"])

hotel_data = {"全部日": flow_allday["hotel_consume_level"], "节假日": flow_holiday["hotel_consume_level"], "周末": flow_weekend["hotel_consume_level"]}
row = write_section(ws1, row, "酒店价格偏好(元)", hotel_data, ["全部日", "节假日", "周末"])

travel_data = {"全部日": flow_allday["travel_tool"], "节假日": flow_holiday["travel_tool"], "周末": flow_weekend["travel_tool"]}
row = write_section(ws1, row, "长途交通方式", travel_data, ["全部日", "节假日", "周末"])

trip_data = {"全部日": flow_allday["trip_mode_r3m"]}
row = write_section(ws1, row, "日常出行交通偏好(近3月)", trip_data, ["全部日"])

for col in range(1, 6):
    ws1.column_dimensions[get_column_letter(col)].width = 16
ws1.column_dimensions['A'].width = 20

# ═══ Sheet 2: 常驻人口画像 ═══
ws2 = wb.create_sheet("常驻人口画像")
ws2.cell(row=1, column=1, value="香洲区(440402)常驻人口画像 - 2026年6月").font = title_font
ws2.cell(row=2, column=1, value="数据来源: 高德云图 cloudmap-gateway | 四种口径: 居住/工作/活跃").font = Font(size=9, name='Microsoft YaHei', color='666666')

# UV汇总
row = 4
ws2.cell(row=row, column=1, value="常驻人口指数(UV)").font = sub_font
row += 1
for col, h in enumerate(["口径", "UV指数", "说明"], 1):
    cell = ws2.cell(row=row, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = thin_border
row += 1
pmnt_uv = [
    ("居住地(home)", pmnt_home["uv"], "在该区有住所的人群"),
    ("工作地(work)", pmnt_work["uv"], "在该区有工作地点的人群"),
    ("活跃地(active)", pmnt_active["uv"], "该月内曾到访该区的人群(含流动人口)"),
]
for typ, uv, desc in pmnt_uv:
    for col, val in enumerate([typ, f"{uv:,}", desc], 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.font = body_font
        cell.border = thin_border
    row += 1

row += 1

# 年龄
age_pmnt = {"居住": pmnt_home["age"], "工作": pmnt_work["age"], "活跃": pmnt_active["age"]}
row = write_section(ws2, row, "年龄分布", age_pmnt, ["居住", "工作", "活跃"])

sex_pmnt = {"居住": pmnt_home["sex"], "工作": pmnt_work["sex"], "活跃": pmnt_active["sex"]}
row = write_section(ws2, row, "性别分布", sex_pmnt, ["居住", "工作", "活跃"])

cons_pmnt = {"居住": pmnt_home["consumption_level"], "工作": pmnt_work["consumption_level"], "活跃": pmnt_active["consumption_level"]}
row = write_section(ws2, row, "消费水平", cons_pmnt, ["居住", "工作", "活跃"])

occ_pmnt = {"居住": pmnt_home["occupation"], "工作": pmnt_work["occupation"], "活跃": pmnt_active["occupation"]}
row = write_section(ws2, row, "职业分布", occ_pmnt, ["居住", "工作", "活跃"])

edu_pmnt = {"居住": pmnt_home["education"], "工作": pmnt_work["education"], "活跃": pmnt_active["education"]}
row = write_section(ws2, row, "学历分布(已知部分,占比极低)", edu_pmnt, ["居住", "工作", "活跃"])

for col in range(1, 6):
    ws2.column_dimensions[get_column_letter(col)].width = 16
ws2.column_dimensions['A'].width = 24

# ═══ Sheet 3: 研究建议 ═══
ws3 = wb.create_sheet("研究解读与建议")
ws3.cell(row=1, column=1, value="桂山岛文旅规划参考解读").font = title_font

notes = [
    ("客流规模", f"香洲区月均日客流约228万(UV指数),节假日略高(231万),桂山镇作为海岛旅游区,实际到岛客流远小于区级数据"),
    ("核心客群年龄", "35-44岁占比最高(约33%),为家庭出游主力;25-34岁青年群体占23%,适合发展海岛露营/水上运动"),
    ("性别比", "男性略多(55.7%),反映海岛旅游偏好男性;但女性44.3%也不低,网红打卡/民宿体验可吸引女性客群"),
    ("消费水平", "中等消费占45.6%,中高占30%;餐饮消费集中在50元以下(58.8%)和50-99元(35.6%),岛上餐饮定价应在此区间"),
    ("住宿偏好", "0-99元(41.7%)和100-299元(42.3%)合计超84%,高端住宿(1000+)仅7.8%;民宿为主力,不宜过度开发高端酒店"),
    ("交通方式", "火车为远途首选(66%),飞机24%;到岛需先从珠海站/金湾机场转船,建议优化码头接驳和船班时刻"),
    ("出行偏好", "日常驾车44.6%为主(本地居民),火车15.8%+打车10.6%(游客);岛上无机动车,步行+骑行体验是关键"),
    ("活跃vs居住", "活跃人口353万远大于居住168万+工作104万,说明香洲区有大量外来流动人群(含游客),是桂山岛潜在客源"),
    ("职业结构", "公司职员49%+商业服务29%为主,白领为绝对主力;可适当开发团建/企业度假产品"),
    ("数据局限", "高德云图最细粒度为区县级(香洲区440402),无法单独剥离桂山镇数据;建议结合船票销售数据交叉验证"),
]

row = 3
for col, h in enumerate(["维度", "解读"], 1):
    cell = ws3.cell(row=row, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = thin_border
row += 1

for dim, note in notes:
    ws3.cell(row=row, column=1, value=dim).font = Font(bold=True, size=10, name='Microsoft YaHei')
    ws3.cell(row=row, column=1).border = thin_border
    ws3.cell(row=row, column=2, value=note).font = body_font
    ws3.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    ws3.cell(row=row, column=2).border = thin_border
    row += 1

ws3.column_dimensions['A'].width = 16
ws3.column_dimensions['B'].width = 80

# ── 保存 ──
OUT = r'C:\Users\Administrator\.qoderwork\workspace\mritsdpuqphp34kx\outputs'
path = os.path.join(OUT, '桂山岛_高德云图分析.xlsx')
wb.save(path)

import shutil
shutil.copy2(path, r'E:\珠海桂山岛案例\数据采集\processed\桂山岛_高德云图分析.xlsx')

print(f"Saved: {path}")
print("Sheet 1: 客流分析 (age/sex/consumption/hotel/travel/trip)")
print("Sheet 2: 常驻人口画像 (home/work/active)")
print("Sheet 3: 研究解读与建议")
