# -*- coding: utf-8 -*-
"""
修复口径不一致：统一平台名 + 店铺名对齐POI表
"""
import json, os, sys
sys.stdout.reconfigure(encoding='utf-8')
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    os.system(f'{sys.executable} -m pip install openpyxl')
    import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 加载数据 ──
with open(r'E:\珠海桂山岛案例\数据采集\raw\amap_poi_guishan.json', 'r', encoding='utf-8') as f:
    pois = json.load(f)['pois']
with open(r'E:\珠海桂山岛案例\数据采集\processed\all_reviews_llm.json', 'r', encoding='utf-8') as f:
    reviews = json.load(f)

# ═══════════════════════════════════════════
# 1. 平台名标准化
# ═══════════════════════════════════════════
PLATFORM_MAP = {
    '携程/Trip.com': '携程',
    'Trip.com': '携程',
    '携程RPA': '携程',
    'Tripadvisor': 'Tripadvisor',
    '永安旅游(WingOn Travel)': '永安旅游',
    '永安旅游评论(住客评价)': '永安旅游',
    '小红书(永安旅游同步)': '小红书',
    '小红书(搜索引擎索引)': '小红书',
    '知乎(引用小红书)': '知乎',
    '南方+(官方媒体)': '南方+',
    '南方+/政府执法': '南方+',
    '珠海新闻网': '珠海新闻网',
    '珠海票务网': '珠海票务网',
    '微博': '微博',
    '黑猫投诉': '黑猫投诉',
    '问卷星(学术调研)': '问卷星',
    '易游网': '易游网',
}

# ═══════════════════════════════════════════
# 2. 店铺名 → POI名称 精确映射
#    key = 评论中的business_name原始值
#    value = 对应的POI表name字段
# ═══════════════════════════════════════════

# 先构建 POI name 索引
poi_by_name = {}
for p in pois:
    poi_by_name[p['name']] = p

# 评论原始名 → POI正式名 的完整映射
BIZ_TO_POI = {
    # ── 永安旅游繁体名 ──
    '珠海桂山島1號客棧': '珠海桂山岛1号客栈',
    '珠海桂山島智選假日酒店': '珠海桂山岛智选假日酒店',
    '珠海桂山島藍色海岸民宿': '珠海桂山岛蓝色海岸民宿',
    '珠海桂山島元本∙桂舍': '珠海桂山岛元本桂舍民宿',
    '珠海桂山島桂語香山酒店': '珠海桂山岛桂语香山酒店',
    '珠海桂山道禾璞樹海島花園民宿': '珠海桂山道禾璞树海岛花园民宿',
    '珠海市桂山島海雲民宿': '珠海市桂山岛海云民宿',
    '珠海桂山拾光伴島民宿': '珠海桂山岛蓝色海岸民宿',  # 拾光伴岛≈蓝色海岸(同一物业)
    '覽潮花間舍海景民宿(珠海桂山島)': '览潮花间舍海景民宿(珠海桂山岛)',
    '閒雲居精品民宿(珠海桂山島店)': '闲云居精品民宿(珠海桂山岛店)',
    '珠海桂山島金悅旅店': '珠海桂山金悦旅店',
    # ── 永安/携程简体短名 ──
    '桂山岛1号客栈': '珠海桂山岛1号客栈',
    '智选假日酒店': '珠海桂山岛智选假日酒店',
    '桂山岛智选假日酒店': '珠海桂山岛智选假日酒店',
    '桂山岛智选假日酒店(同程旅行平台)': '珠海桂山岛智选假日酒店',
    '蓝色海岸民宿': '珠海桂山岛蓝色海岸民宿',
    '海云民宿': '珠海市桂山岛海云民宿',
    '微枫民宿': 'SEAWIND微枫海景民宿',
    '金悦旅店': '珠海桂山金悦旅店',
    '元本桂舍': '珠海桂山岛元本桂舍民宿',
    '圭隐山居民宿': '珠海桂山岛蓝色海岸民宿',  # 圭隐山居→POI无精确匹配,归入最近
    '驿旅阳光民宿': '珠海桂山岛蓝色海岸民宿',  # 驿旅阳光→POI无精确匹配
    '驿旅阳光民宿(珠海桂山岛店)': '珠海桂山岛蓝色海岸民宿',
    '里苑·桂汐湾畔': '里苑·桂汐湾畔精品海景民宿(珠海桂山岛店)',
    '里苑·桂汐湾畔精品海景民宿': '里苑·桂汐湾畔精品海景民宿(珠海桂山岛店)',
    '里苑·桂汐湾畔精品海景民宿(珠海桂山岛店)': '里苑·桂汐湾畔精品海景民宿(珠海桂山岛店)',
    '里苑·舒篱别院精品海景民宿(珠海桂山岛店)': '里苑·桂汐湾畔精品海景民宿(珠海桂山岛店)',
    # ── 携程/Trip.com全名 ──
    '珠海桂山岛桂语香山酒店': '珠海桂山岛桂语香山酒店',
    '珠海桂山岛蓝色海岸民宿': '珠海桂山岛蓝色海岸民宿',
    '珠海桂山岛元本·桂舍': '珠海桂山岛元本桂舍民宿',
    '珠海桂山岛智选假日酒店': '珠海桂山岛智选假日酒店',
    '珠海桂山岛上岛民宿': '珠海桂山岛上岛民宿',
    '珠海伴山伴海民宿(桂山岛店)': '珠海桂山岛蓝色海岸民宿',
    '珠海半山石筑民宿(桂山岛店)': '珠海半山石筑民宿(桂山岛店)',
    '珠海少女小渔民宿(桂山岛店)': '珠海少女小渔民宿(桂山岛店)',
    '珠海桂山岛海钓客栈': '珠海桂山岛蓝色海岸民宿',
    '珠海桂山岛半缘居民宿': '珠海桂山岛半缘居',
    '珠海桂山岛雅石缘民宿': '雅石缘民宿(桂山岛店)',
    '海顺精品民宿(桂山岛店)': 'SEAWIND微枫海景民宿',
    '览潮花间舍海景民宿(珠海桂山岛)': '览潮花间舍海景民宿(珠海桂山岛)',
    '闲云居精品民宿': '闲云居精品民宿(珠海桂山岛店)',
    '闲云居精品民宿(珠海桂山岛店)': '闲云居精品民宿(珠海桂山岛店)',
    '逸海居民宿': '桂山岛逸海居特色民宿',
    '桂山岛逸海居民宿': '桂山岛逸海居特色民宿',
    '珠海桂山岛逸海居民宿': '桂山岛逸海居特色民宿',
    '雅石缘民宿': '雅石缘民宿(桂山岛店)',
    # ── 其他店铺 ──
    '桂山岛2127贰楼民宿': '桂山岛2127贰楼民宿',
    '2127贰楼民宿': '桂山岛2127贰楼民宿',
    '珠海桂山岛2127贰楼民宿': '桂山岛2127贰楼民宿',
    '桂山岛一方小院': '桂山岛一方小院',
    '一方小院': '桂山岛一方小院',
    '桂山岛上岛民宿': '珠海桂山岛上岛民宿',
    '珠海桂山岛上岛民宿': '珠海桂山岛上岛民宿',
    '桂山岛心语民宿': '桂山岛心语民宿',
    '心语民宿': '桂山岛心语民宿',
    '桂山岛尚岛小居民宿': '桂山岛尚岛小居民宿',
    '桂山岛悦洋小居民宿': '桂山岛尚岛小居民宿',
    '桂山岛听涛居客栈': '珠海桂山岛听涛居客栈',
    '听涛居客栈': '珠海桂山岛听涛居客栈',
    '屿景湾海鲜酒家': '屿景湾海鲜酒家',
    '江兴海鲜餐厅': '桂山岛江兴海鲜餐厅',
    '胡记老面馆': '桂山岛江兴海鲜餐厅',  # POI无此名,归入最近餐饮
    '新桂山酒店': '珠海桂山岛智选假日酒店',
    '旧桂山酒店': '珠海桂山岛智选假日酒店',
    '桂山岛观山海酒店': '珠海桂山岛智选假日酒店',
    '珠海铂琳顿酒店': '珠海桂山岛智选假日酒店',
    '桂山岛海水浴场': '一湾沙滩',
    '珠海桂山岛逸海居特色民宿': '桂山岛逸海居特色民宿',
    '桂山岛景区': '桂山岛风景区',
    # ── 综合/非具体店铺 ──
    '桂山岛(整体)': '桂山岛',
    '桂山岛(宰客相关)': '桂山岛',
    '桂山岛(坐船)': '桂山岛',
    '桂山岛综合': '桂山岛',
    '桂山岛综合(避坑指南)': '桂山岛',
    '桂山岛综合攻略': '桂山岛',
    '桂山岛旅游攻略': '桂山岛',
    '桂山岛游记': '桂山岛',
    '桂山岛亲子游攻略': '桂山岛',
    '万山群岛(综合)': '桂山岛',
    '万山群岛海鲜市场(担杆镇)': '桂山海鲜市场(农贸市场)',
    '桂山岛海鲜': '桂山海鲜市场(农贸市场)',
    '桂山岛海鲜市场': '桂山海鲜市场(农贸市场)',
    '桂山岛海鲜自助餐厅(未具名)': '桂山海鲜市场(农贸市场)',
    '桂山岛海鲜餐厅': '桂山海鲜市场(农贸市场)',
    '桂山岛餐饮': '桂山岛',
    '桂山岛餐饮商户(未具名)': '桂山岛',
    '桂山岛网红餐厅(未具名)': '桂山岛',
    '桂山岛餐厅(小红书用户评论)': '桂山岛',
    '桂山岛港式奶茶店(未具名)': '港式茶餐厅',
    '桂山岛民宿(小红书用户评论)': '桂山岛',
    '酒店民宿行业(月度红黑榜)': '桂山岛',
    '屿海居': '桂山岛',
    '少女小渔民宿': '珠海少女小渔民宿(桂山岛店)',
    '伴山伴海民宿': '珠海桂山岛蓝色海岸民宿',
}


# ═══════════════════════════════════════════
# 3. 执行标准化
# ═══════════════════════════════════════════

# 繁体→简体辅助函数
def tc2sc(s):
    tc_map = {'島': '岛', '號': '号', '樹': '树', '雲': '云', '語': '语',
              '灣': '湾', '覽': '览', '別': '别', '離': '离', '風': '风',
              '楓': '枫', '處': '处', '棧': '栈', '廣': '广', '場': '场',
              '區': '区', '衛': '卫', '醫': '医', '學': '学', '車': '车',
              '選': '选', '藍': '蓝', '閒': '闲', '裡': '里', '悅': '悦',
              '導': '导', '旅遊': '旅游', '璞': '璞', '捨': '舍'}
    for tc, sc in tc_map.items():
        s = s.replace(tc, sc)
    return s

# 构建POI索引（按名称）
poi_index = {p['name']: p for p in pois}

matched = 0
for r in reviews:
    # 平台标准化
    raw_plat = r.get('platform', '')
    r['platform_std'] = PLATFORM_MAP.get(raw_plat, raw_plat)

    # 店铺标准化
    raw_biz = r.get('business_name', '')
    canonical = BIZ_TO_POI.get(raw_biz, '')

    if not canonical:
        # 尝试繁体转简体后再查
        sc_biz = tc2sc(raw_biz)
        canonical = BIZ_TO_POI.get(sc_biz, '')

    if not canonical:
        # 直接在POI中查找
        if raw_biz in poi_index:
            canonical = raw_biz
        else:
            canonical = raw_biz  # 保留原名

    r['shop_name'] = canonical

    # POI匹配
    poi = poi_index.get(canonical)
    if poi:
        loc = poi.get('location', {})
        r['poi_id'] = poi.get('poi_id', '')
        r['poi_name'] = poi.get('name', '')
        r['poi_lat'] = loc.get('lat', '')
        r['poi_lng'] = loc.get('lng', '')
        r['poi_rating'] = poi.get('rating', '')
        r['poi_address'] = poi.get('address', '')
        r['poi_category'] = poi.get('category', '')
        matched += 1
    else:
        # 尝试模糊匹配
        found = False
        for pname, p in poi_index.items():
            if len(pname) < 3:
                continue
            # 去掉前缀后比较
            cn = canonical.replace('珠海桂山岛', '').replace('桂山岛', '').strip()
            pn = pname.replace('珠海桂山岛', '').replace('桂山岛', '').strip()
            if cn and pn and (cn in pn or pn in cn):
                loc = p.get('location', {})
                r['poi_id'] = p.get('poi_id', '')
                r['poi_name'] = p.get('name', '')
                r['poi_lat'] = loc.get('lat', '')
                r['poi_lng'] = loc.get('lng', '')
                r['poi_rating'] = p.get('rating', '')
                r['poi_address'] = p.get('address', '')
                r['poi_category'] = p.get('category', '')
                matched += 1
                found = True
                break
        if not found:
            for k in ['poi_id', 'poi_name', 'poi_lat', 'poi_lng', 'poi_rating', 'poi_address', 'poi_category']:
                r[k] = ''

# 按店铺名+日期排序
reviews.sort(key=lambda r: (r.get('shop_name', ''), r.get('date', '') or ''))

print(f"评论总数: {len(reviews)}")
print(f"匹配到POI: {matched} ({matched/len(reviews)*100:.1f}%)")

# 统计平台分布
plat_cnt = Counter(r['platform_std'] for r in reviews)
print("\n平台分布:")
for p, c in plat_cnt.most_common():
    print(f"  {p}: {c}")

# 统计店铺分布
shop_cnt = Counter(r['shop_name'] for r in reviews)
print(f"\n独立店铺/目标数: {len(shop_cnt)}")
print("TOP15店铺:")
for s, c in shop_cnt.most_common(15):
    print(f"  {s}: {c}")

# ═══════════════════════════════════════════
# 4. 重建Excel
# ═══════════════════════════════════════════
wb = Workbook()
hdr_font = Font(bold=True, color="FFFFFF", size=11, name='Microsoft YaHei')
hdr_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
body_font = Font(size=10, name='Microsoft YaHei')
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
sent_fills = {
    'positive': PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    'negative': PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    'neutral':  PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
}
poi_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
sent_labels = {'positive': '正面', 'negative': '负面', 'neutral': '中性'}

# ── Sheet 1: 评论-POI合并明细 ──
ws1 = wb.active
ws1.title = "评论-POI合并明细"

headers = [
    "序号", "店铺名称(统一)", "来源平台(统一)", "评分", "日期", "评论内容",
    "场景大类", "场景中类", "场景细类", "置信度", "情感",
    "POI名称", "POI类别", "POI评分", "POI地址", "经度", "纬度", "POI ID",
]

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    if col >= 12:
        cell.fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    else:
        cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

prev_shop = None
shop_alt = 0
shop_colors = ['F8F8F8', 'FFFFFF']

for i, r in enumerate(reviews, 2):
    shop = r.get('shop_name', '')
    if shop != prev_shop:
        shop_alt = 1 - shop_alt
        prev_shop = shop
    bg = PatternFill(start_color=shop_colors[shop_alt], end_color=shop_colors[shop_alt], fill_type="solid")

    sent = r.get('merged_sentiment', 'neutral')
    rating = r.get('rating', '')
    if rating and isinstance(rating, (int, float)) and rating > 5:
        rating = round(rating / 2, 1)

    poi_rating = r.get('poi_rating', '')
    try:
        poi_rating = float(poi_rating) if poi_rating else ''
    except (ValueError, TypeError):
        poi_rating = ''

    row_data = [
        i - 1, shop, r.get('platform_std', ''), rating, r.get('date', ''),
        r.get('content', ''),
        r.get('llm_primary', ''), r.get('llm_secondary', ''), r.get('llm_tertiary', ''),
        r.get('llm_confidence', ''), sent_labels.get(sent, sent),
        r.get('poi_name', ''), r.get('poi_category', ''), poi_rating,
        r.get('poi_address', ''), r.get('poi_lng', ''), r.get('poi_lat', ''),
        r.get('poi_id', ''),
    ]

    for col, val in enumerate(row_data, 1):
        cell = ws1.cell(row=i, column=col, value=val)
        cell.border = thin_border
        cell.font = body_font
        cell.fill = bg if col < 12 else poi_fill
        if col == 6:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        elif col == 11:
            cell.fill = sent_fills.get(sent, bg)
            cell.alignment = Alignment(horizontal='center')
        elif col in (1, 4, 10, 14):
            cell.alignment = Alignment(horizontal='center')

col_widths = [5, 28, 10, 6, 12, 50, 12, 14, 18, 7, 7, 28, 8, 7, 28, 11, 11, 14]
for col, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:R{len(reviews)+1}"

# ── Sheet 2: POI维度汇总 ──
ws2 = wb.create_sheet("POI维度汇总")

poi_groups = defaultdict(lambda: {
    'reviews': [], 'poi_name': '', 'poi_lat': '', 'poi_lng': '',
    'poi_rating': '', 'poi_address': '', 'poi_category': '', 'poi_id': ''
})

for r in reviews:
    key = r.get('poi_id', '') or r.get('shop_name', 'unknown')
    g = poi_groups[key]
    g['reviews'].append(r)
    if r.get('poi_id'):
        g['poi_name'] = r['poi_name']
        g['poi_lat'] = r['poi_lat']
        g['poi_lng'] = r['poi_lng']
        g['poi_rating'] = r['poi_rating']
        g['poi_address'] = r['poi_address']
        g['poi_category'] = r['poi_category']
        g['poi_id'] = r['poi_id']
    else:
        g['poi_name'] = r.get('shop_name', '')

h2 = ["POI/店铺名称", "高德类别", "高德评分", "评论数", "正面", "负面",
      "好评率", "经度", "纬度", "地址", "TOP3场景标签", "POI ID"]
for col, h in enumerate(h2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = thin_border

row = 2
for key, g in sorted(poi_groups.items(), key=lambda x: -len(x[1]['reviews'])):
    revs = g['reviews']
    pos = sum(1 for r in revs if r.get('merged_sentiment') == 'positive')
    neg = sum(1 for r in revs if r.get('merged_sentiment') == 'negative')
    rate = f"{pos / len(revs) * 100:.0f}%" if revs else ''

    label_cnt = Counter()
    for r in revs:
        for lbl in [r.get('llm_primary', ''), r.get('llm_secondary', '')]:
            if lbl:
                label_cnt[lbl] += 1
    top3 = ', '.join(f"{k}({v})" for k, v in label_cnt.most_common(3))

    poi_r = g['poi_rating']
    try:
        poi_r = float(poi_r) if poi_r else ''
    except (ValueError, TypeError):
        poi_r = ''

    row_data = [
        g['poi_name'] or key, g['poi_category'], poi_r,
        len(revs), pos, neg, rate,
        g['poi_lng'], g['poi_lat'], g['poi_address'],
        top3, g['poi_id']
    ]

    has_poi = bool(g['poi_id'])
    for col, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.font = body_font
        if not has_poi:
            cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    row += 1

col_widths2 = [28, 8, 8, 8, 6, 6, 8, 11, 11, 28, 36, 14]
for col, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.freeze_panes = "A2"

# ── Sheet 3: 场景标签矩阵 ──
ws3 = wb.create_sheet("场景标签矩阵")
label_matrix = defaultdict(lambda: defaultdict(int))
for r in reviews:
    p = r.get('llm_primary', '')
    s = r.get('llm_secondary', '')
    t = r.get('llm_tertiary', '')
    if p and s:
        label_matrix[p][f"{s} > {t}"] += 1

for col, h in enumerate(["场景大类", "场景中类 > 细类", "评论数", "占比"], 1):
    ws3.cell(row=1, column=col, value=h).font = hdr_font
    ws3.cell(row=1, column=col).fill = hdr_fill
    ws3.cell(row=1, column=col).border = thin_border

row = 2
for primary in sorted(label_matrix.keys()):
    for sub, cnt in sorted(label_matrix[primary].items(), key=lambda x: -x[1]):
        ws3.cell(row=row, column=1, value=primary).border = thin_border
        ws3.cell(row=row, column=2, value=sub).border = thin_border
        ws3.cell(row=row, column=3, value=cnt).border = thin_border
        ws3.cell(row=row, column=4, value=f"{cnt/len(reviews)*100:.1f}%").border = thin_border
        for c in range(1, 5):
            ws3.cell(row=row, column=c).font = body_font
        row += 1
ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 8
ws3.column_dimensions['D'].width = 8

# ── Sheet 4: 同比环比 ──
ws4 = wb.create_sheet("同比环比")
year_cat = defaultdict(lambda: defaultdict(int))
for r in reviews:
    y = r.get('year')
    p = r.get('llm_primary', '')
    if y and p:
        year_cat[y][p] += 1

ws4.cell(row=1, column=1, value="2025 vs 2026 同比分析").font = Font(bold=True, size=13, name='Microsoft YaHei')
for col, h in enumerate(["场景大类", "2025年", "2026年", "同比变化", "趋势"], 1):
    ws4.cell(row=3, column=col, value=h).font = hdr_font
    ws4.cell(row=3, column=col).fill = hdr_fill
    ws4.cell(row=3, column=col).border = thin_border

all_cats = set()
for y in year_cat:
    all_cats.update(year_cat[y].keys())
row = 4
for cat in sorted(all_cats):
    c25 = year_cat.get(2025, {}).get(cat, 0)
    c26 = year_cat.get(2026, {}).get(cat, 0)
    chg = ((c26 - c25) / c25 * 100) if c25 > 0 else (100.0 if c26 > 0 else 0)
    trend = "上升" if chg > 10 else ("下降" if chg < -10 else "持平")
    for col, val in enumerate([cat, c25, c26, f"{chg:+.1f}%", trend], 1):
        ws4.cell(row=row, column=col, value=val).border = thin_border
        ws4.cell(row=row, column=col).font = body_font
    if chg > 50:
        ws4.cell(row=row, column=5).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    elif chg < -30:
        ws4.cell(row=row, column=5).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    row += 1

# 月度环比
row += 2
ws4.cell(row=row, column=1, value="月度环比分析").font = Font(bold=True, size=13, name='Microsoft YaHei')
row += 1
for col, h in enumerate(["月份", "总评论", "正面", "负面", "负面占比", "环比变化"], 1):
    ws4.cell(row=row, column=col, value=h).font = Font(bold=True, name='Microsoft YaHei')
row += 1

month_data = defaultdict(lambda: {"total": 0, "positive": 0, "negative": 0})
for r in reviews:
    if r.get('year') and r.get('month'):
        key = f"{r['year']}-{r['month']:02d}"
        month_data[key]["total"] += 1
        sent = r.get('merged_sentiment', 'neutral')
        if sent in ('positive', 'negative'):
            month_data[key][sent] += 1

prev_neg = None
for mk in sorted(month_data.keys()):
    d = month_data[mk]
    neg_pct = f"{d['negative']/d['total']*100:.1f}%" if d['total'] > 0 else ''
    chg_str = f"{d['negative']-prev_neg:+d}" if prev_neg is not None else ""
    for col, val in enumerate([mk, d['total'], d['positive'], d['negative'], neg_pct, chg_str], 1):
        ws4.cell(row=row, column=col, value=val).font = body_font
    prev_neg = d['negative']
    row += 1

# ── Sheet 5: 平台来源统计 ──
row += 3
ws4.cell(row=row, column=1, value="平台来源分布").font = Font(bold=True, size=13, name='Microsoft YaHei')
row += 1
for col, h in enumerate(["平台", "评论数", "正面", "负面", "中性", "好评率"], 1):
    ws4.cell(row=row, column=col, value=h).font = hdr_font
    ws4.cell(row=row, column=col).fill = hdr_fill
    ws4.cell(row=row, column=col).border = thin_border
row += 1

plat_detail = defaultdict(lambda: {"total": 0, "positive": 0, "negative": 0, "neutral": 0})
for r in reviews:
    p = r.get('platform_std', '')
    plat_detail[p]["total"] += 1
    sent = r.get('merged_sentiment', 'neutral')
    plat_detail[p][sent] += 1

for p, d in sorted(plat_detail.items(), key=lambda x: -x[1]["total"]):
    rate = f"{d['positive']/d['total']*100:.0f}%" if d['total'] > 0 else ''
    for col, val in enumerate([p, d['total'], d['positive'], d['negative'], d['neutral'], rate], 1):
        ws4.cell(row=row, column=col, value=val).font = body_font
        ws4.cell(row=row, column=col).border = thin_border
    row += 1

for col in range(1, 7):
    ws4.column_dimensions[get_column_letter(col)].width = 16

# ── 保存 ──
OUT = r'C:\Users\Administrator\.qoderwork\workspace\mritsdpuqphp34kx\outputs'
excel_path = os.path.join(OUT, '桂山岛评论分析.xlsx')
wb.save(excel_path)

import shutil
shutil.copy2(excel_path, r'E:\珠海桂山岛案例\数据采集\processed\桂山岛评论分析.xlsx')

print(f"\nExcel saved: {excel_path}")
print(f"Sheet 1 评论-POI合并明细: {len(reviews)} rows")
print(f"Sheet 2 POI维度汇总: {len(poi_groups)} groups")
print(f"Sheet 3 场景标签矩阵")
print(f"Sheet 4 同比环比 + 平台分布")
