# -*- coding: utf-8 -*-
"""导出桂山岛AOI级客流+常驻数据，与香洲区对比"""
import json, os, sys
sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
except ImportError:
    os.system(f'{sys.executable} -m pip install openpyxl')
    import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = load_workbook(r'E:\珠海桂山岛案例\数据采集\processed\桂山岛_高德云图分析.xlsx')

hdr_font = Font(bold=True, color="FFFFFF", size=11, name='Microsoft YaHei')
hdr_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
body_font = Font(size=10, name='Microsoft YaHei')
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
title_font = Font(bold=True, size=13, name='Microsoft YaHei')
sub_font = Font(bold=True, size=11, name='Microsoft YaHei', color='1A2744')
hl_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

# ═══ AOI数据 ═══
aoi_allday = {
    "uv": 4994,
    "age": {"19-24": 0.042797, "25-29": 0.111292, "30-34": 0.110197, "35-39": 0.154197,
            "40-44": 0.151693, "45-49": 0.141141, "50-54": 0.144386, "55-59": 0.091461,
            "60-64": 0.038774, "65+": 0.014062},
    "sex": {"男": 0.660051, "女": 0.339949},
    "consumption_level": {"低": 0.188585, "中": 0.455537, "中高": 0.266645, "高": 0.089233},
    "dining_consum_level": {"0-49": 0.544222, "50-99": 0.411096, "100-199": 0.039936,
                            "200-299": 0.002840, "300-399": 0.001157, "400-499": 0.000357, "500+": 0.000391},
    "hotel_consume_level": {"0-99": 0.487947, "100-299": 0.410494, "300-499": 0.034231,
                            "500-999": 0.038909, "1000+": 0.028419},
    "travel_tool": {"火车": 0.583172, "飞机": 0.302394, "汽车": 0.114434},
    "trip_mode_r3m": {"驾车": 0.348162, "地铁": 0.196021, "火车": 0.146529, "打车": 0.099215,
                      "飞机": 0.064696, "公交": 0.051296, "步行": 0.050702, "骑行": 0.036696,
                      "摩托": 0.004948, "长途汽车": 0.001735},
    "hotel_type_favorite": {"四星级宾馆": 0.273263, "旅馆招待所": 0.230958, "五星级宾馆": 0.226737,
                            "三星级宾馆": 0.145958, "经济型连锁酒店": 0.117814, "青年旅舍": 0.004641, "奢华酒店": 0.000629},
}
aoi_weekend = {
    "uv": 5637,
    "age": {"19-24": 0.044163, "25-29": 0.128780, "30-34": 0.128000, "35-39": 0.160943,
            "40-44": 0.149951, "45-49": 0.133756, "50-54": 0.124683, "55-59": 0.082439,
            "60-64": 0.034862, "65+": 0.012423},
    "sex": {"男": 0.626728, "女": 0.373272},
    "consumption_level": {"低": 0.176151, "中": 0.437505, "中高": 0.278089, "高": 0.108255},
    "dining_consum_level": {"0-49": 0.552133, "50-99": 0.401692, "100-199": 0.040603,
                            "200-299": 0.003147, "300-399": 0.001341, "400-499": 0.000464, "500+": 0.000619},
    "hotel_consume_level": {"0-99": 0.484106, "100-299": 0.401489, "300-499": 0.036813,
                            "500-999": 0.043170, "1000+": 0.034421},
    "travel_tool": {"火车": 0.572943, "飞机": 0.307436, "汽车": 0.119621},
}
aoi_holiday = {
    "uv": 5620,
    "age": {"19-24": 0.044716, "25-29": 0.130368, "30-34": 0.127460, "35-39": 0.160110,
            "40-44": 0.150196, "45-49": 0.133246, "50-54": 0.125134, "55-59": 0.081582,
            "60-64": 0.034947, "65+": 0.012240},
    "sex": {"男": 0.625009, "女": 0.374991},
    "consumption_level": {"低": 0.177293, "中": 0.436278, "中高": 0.278140, "高": 0.108289},
    "travel_tool": {"火车": 0.573427, "飞机": 0.307118, "汽车": 0.119455},
}
aoi_active = {
    "uv": 7462,
    "age": {"19-24": 0.042404, "25-29": 0.087359, "30-34": 0.098996, "35-39": 0.150964,
            "40-44": 0.151921, "45-49": 0.149370, "50-54": 0.157500, "55-59": 0.097561,
            "60-64": 0.046070, "65+": 0.017854},
    "sex": {"男": 0.672039, "女": 0.327961},
    "consumption_level": {"低": 0.200482, "中": 0.486197, "中高": 0.251139, "高": 0.062182},
    "occupation": {"商业服务职员": 0.390925, "公司职员": 0.384817, "其他": 0.145724,
                   "医护人员": 0.046248, "生产运输职员": 0.023560, "教职工": 0.008726},
    "education": {"本科": 0.003886, "专科": 0.001340, "硕士": 0.000938},
}
aoi_home = {
    "uv": 1605,
    "age": {"19-24": 0.039552, "25-29": 0.094776, "30-34": 0.096269, "35-39": 0.131343,
            "40-44": 0.158955, "45-49": 0.144776, "50-54": 0.165672, "55-59": 0.121642,
            "60-64": 0.037313, "65+": 0.009701},
    "sex": {"男": 0.648393, "女": 0.351607},
    "consumption_level": {"低": 0.333333, "中": 0.541433, "中高": 0.109657, "高": 0.015576},
    "occupation": {"商业服务职员": 0.463504, "公司职员": 0.343066, "其他": 0.124088,
                   "医护人员": 0.051095, "生产运输职员": 0.018248},
}

# 香洲区对照数据
dist_allday = {
    "uv": 2284077,
    "age": {"19-24": 0.055584, "25-29": 0.103054, "30-34": 0.123390, "35-39": 0.171978,
            "40-44": 0.160421, "45-49": 0.133867, "50-54": 0.116163, "55-59": 0.075949,
            "60-64": 0.038170, "65+": 0.021425},
    "sex": {"男": 0.557181, "女": 0.442819},
    "consumption_level": {"低": 0.115552, "中": 0.456174, "中高": 0.299882, "高": 0.128392},
    "travel_tool": {"火车": 0.659788, "飞机": 0.244461, "汽车": 0.095751},
}

def write_compare(ws, start_row, title, label_order, aoi_data, dist_data, fmt='pct'):
    row = start_row
    ws.cell(row=row, column=1, value=title).font = sub_font
    row += 1
    headers = ["类别", "桂山岛AOI", "香洲区(对照)", "差异"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin_border
    row += 1
    for label in label_order:
        aoi_v = aoi_data.get(label, 0)
        dist_v = dist_data.get(label, 0)
        diff = aoi_v - dist_v
        ws.cell(row=row, column=1, value=label).font = body_font
        ws.cell(row=row, column=1).border = thin_border
        for col, val in enumerate([f"{aoi_v*100:.2f}%", f"{dist_v*100:.2f}%", f"{diff*100:+.2f}pp"], 2):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = body_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        # 高亮显著差异
        if abs(diff) > 0.05:
            ws.cell(row=row, column=4).fill = hl_fill
        row += 1
    return row + 1

# ═══ 新建Sheet: AOI级分析 ═══
ws = wb.create_sheet("桂山岛AOI级分析", 0)  # 插到最前面

ws.cell(row=1, column=1, value="桂山岛风景区(AOI)客流画像 vs 香洲区(行政区)对照").font = title_font
ws.cell(row=2, column=1, value="AOI ID: B0FFFQ7RVA (桂山岛风景区,4A) | 2026年6月 | 数据来源: 高德云图").font = Font(size=9, name='Microsoft YaHei', color='666666')

# UV对比
row = 4
ws.cell(row=row, column=1, value="客流规模对比(UV指数)").font = sub_font
row += 1
for col, h in enumerate(["类型", "桂山岛AOI", "香洲区", "AOI占比"], 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = thin_border
row += 1
uv_pairs = [
    ("全部日(月均)", 4994, 2284077),
    ("周末(日均)", 5637, 2313960),
    ("节假日(日均)", 5620, 2316671),
]
for typ, aoi, dist in uv_pairs:
    ratio = aoi / dist * 100
    for col, val in enumerate([typ, f"{aoi:,}", f"{dist:,}", f"{ratio:.3f}%"], 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = body_font
        cell.border = thin_border
        if col >= 2:
            cell.alignment = Alignment(horizontal='center')
    row += 1

row += 1

# 年龄对比
age_labels = ["19-24", "25-29", "30-34", "35-39", "40-44", "45-49", "50-54", "55-59", "60-64", "65+"]
row = write_compare(ws, row, "年龄分布对比", age_labels, aoi_allday["age"], dist_allday["age"])

# 性别对比
row = write_compare(ws, row, "性别分布对比", ["男", "女"], aoi_allday["sex"], dist_allday["sex"])

# 消费水平对比
row = write_compare(ws, row, "消费水平对比", ["低", "中", "中高", "高"], aoi_allday["consumption_level"], dist_allday["consumption_level"])

# 交通方式对比
row = write_compare(ws, row, "长途交通方式对比", ["火车", "飞机", "汽车"], aoi_allday["travel_tool"], dist_allday["travel_tool"])

# AOI独有数据: 酒店类型偏好
ws.cell(row=row, column=1, value="酒店类型偏好(桂山岛AOI独有)").font = sub_font
row += 1
for col, h in enumerate(["酒店类型", "占比"], 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = thin_border
row += 1
for ht, pct in sorted(aoi_allday["hotel_type_favorite"].items(), key=lambda x: -x[1]):
    ws.cell(row=row, column=1, value=ht).font = body_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=f"{pct*100:.2f}%").font = body_font
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=2).border = thin_border
    row += 1

row += 1

# 常驻人口(AOI vs 行政区)
ws.cell(row=row, column=1, value="常驻人口画像(AOI级)").font = sub_font
row += 1
for col, h in enumerate(["口径", "桂山岛AOI UV", "说明"], 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = thin_border
row += 1
pmnt_pairs = [
    ("活跃人口(active)", aoi_active["uv"], f"月内到访桂山岛的人群(含游客)"),
    ("居住人口(home)", aoi_home["uv"], f"居住在桂山岛的常住人口"),
]
for typ, uv, desc in pmnt_pairs:
    for col, val in enumerate([typ, f"{uv:,}", desc], 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = body_font
        cell.border = thin_border
    row += 1

row += 1

# 岛民年龄
ws.cell(row=row, column=1, value="岛民(居住)年龄分布").font = sub_font
row += 1
for col, h in enumerate(["年龄段", "岛民", "游客(全部日)", "差异"], 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = thin_border
row += 1
for label in age_labels:
    home_v = aoi_home["age"].get(label, 0)
    flow_v = aoi_allday["age"].get(label, 0)
    diff = home_v - flow_v
    for col, val in enumerate([label, f"{home_v*100:.2f}%", f"{flow_v*100:.2f}%", f"{diff*100:+.2f}pp"], 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = body_font
        cell.border = thin_border
        if col >= 2:
            cell.alignment = Alignment(horizontal='center')
    if abs(diff) > 0.03:
        ws.cell(row=row, column=4).fill = hl_fill
    row += 1

row += 1

# 岛民职业
ws.cell(row=row, column=1, value="岛民(居住)职业分布").font = sub_font
row += 1
for col, h in enumerate(["职业", "岛民", "活跃到访"], 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = thin_border
row += 1
all_occ = set(list(aoi_home["occupation"].keys()) + list(aoi_active["occupation"].keys()))
for occ in sorted(all_occ, key=lambda x: -(aoi_home["occupation"].get(x, 0))):
    hv = aoi_home["occupation"].get(occ, 0)
    av = aoi_active["occupation"].get(occ, 0)
    for col, val in enumerate([occ, f"{hv*100:.2f}%", f"{av*100:.2f}%"], 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = body_font
        cell.border = thin_border
        if col >= 2:
            cell.alignment = Alignment(horizontal='center')
    row += 1

# 消费水平对比(岛民vs游客)
row += 1
ws.cell(row=row, column=1, value="消费水平: 岛民 vs 游客").font = sub_font
row += 1
for col, h in enumerate(["级别", "岛民(居住)", "游客(全部日)", "差异"], 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = thin_border
row += 1
for lvl in ["低", "中", "中高", "高"]:
    hv = aoi_home["consumption_level"].get(lvl, 0)
    fv = aoi_allday["consumption_level"].get(lvl, 0)
    diff = hv - fv
    for col, val in enumerate([lvl, f"{hv*100:.2f}%", f"{fv*100:.2f}%", f"{diff*100:+.2f}pp"], 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = body_font
        cell.border = thin_border
        if col >= 2:
            cell.alignment = Alignment(horizontal='center')
    row += 1

for col in range(1, 6):
    ws.column_dimensions[get_column_letter(col)].width = 18
ws.column_dimensions['A'].width = 22

# ── 保存 ──
OUT = r'C:\Users\Administrator\.qoderwork\workspace\mritsdpuqphp34kx\outputs'
path = os.path.join(OUT, '桂山岛_高德云图分析_v2.xlsx')
wb.save(path)

import shutil
shutil.copy2(path, r'E:\珠海桂山岛案例\数据采集\processed\桂山岛_高德云图分析_v2.xlsx')

print(f"Saved: {path}")
print("New Sheet 1: 桂山岛AOI级分析 (vs 香洲区对照)")
print("  UV: 4994(全部日) / 5637(周末) / 5620(节假日)")
print("  居住人口: 1605 | 活跃人口: 7462")
print("  男性66%(远高于区级55.7%)")
print("  低消费占18.9%(区级11.6%)，飞机30.2%(区级24.4%)")
