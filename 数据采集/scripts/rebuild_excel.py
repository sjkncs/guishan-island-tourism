# -*- coding: utf-8 -*-
"""
重建评论分析Excel：评论-店铺-场景标签一一对应
"""
import json, os, sys
sys.stdout.reconfigure(encoding='utf-8')

from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    os.system(f'{sys.executable} -m pip install openpyxl')
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── 加载数据 ──
with open(r'E:\珠海桂山岛案例\数据采集\processed\all_reviews_llm.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# 标准化商家名（合并同一家）
name_map = {
    '驿旅阳光民宿(珠海桂山岛店)': '驿旅阳光民宿',
    '珠海桂山岛上岛民宿': '上岛民宿',
    '珠海桂山岛蓝色海岸民宿': '蓝色海岸民宿',
    '珠海桂山島智選假日酒店': '智选假日酒店',
    '里苑·桂汐湾畔精品海景民宿(珠海桂山岛店)': '里苑·桂汐湾畔民宿',
    '里苑·桂汐湾畔精品海景民宿': '里苑·桂汐湾畔民宿',
    '珠海桂山岛元本·桂舍': '元本桂舍',
    '海顺精品民宿(桂山岛店)': '海顺精品民宿',
    '览潮花间舍海景民宿(珠海桂山岛)': '览潮花间舍民宿',
    '里苑·舒篱别院精品海景民宿(珠海桂山岛店)': '里苑·舒篱别院民宿',
    '珠海市桂山岛海云民宿': '海云民宿',
    '珠海SEA&WIND微枫民宿(桂山岛店)': '微枫民宿',
    '珠海桂山岛听涛居客栈': '听涛居客栈',
    '珠海桂山岛雅石缘民宿': '雅石缘民宿',
    '珠海桂山岛2127贰楼民宿': '2127贰楼民宿',
    '珠海伴山伴海民宿(桂山岛店)': '伴山伴海民宿',
    '珠海少女小渔民宿(桂山岛店)': '少女小渔民宿',
    '珠海桂山岛逸海居特色民宿': '逸海居民宿',
    '桂山岛一方小院': '一方小院',
    '桂山岛心语民宿': '心语民宿',
    '桂山岛尚岛小居民宿': '尚岛小居民宿',
}

for r in data:
    raw = r.get('business_name', '')
    r['shop_name'] = name_map.get(raw, raw)

# 按商家排序
data.sort(key=lambda r: (r.get('shop_name', ''), r.get('date', '') or ''))

print(f"总评论: {len(data)}")
shops = Counter(r['shop_name'] for r in data)
print(f"标准化后商家数: {len(shops)}")

# ═══ 创建Excel ═══
wb = Workbook()

# ── 样式定义 ──
hdr_font = Font(bold=True, color="FFFFFF", size=11, name='Microsoft YaHei')
hdr_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
body_font = Font(size=10, name='Microsoft YaHei')
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)
sent_fills = {
    'positive': PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    'negative': PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    'neutral':  PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
}
sent_labels = {'positive': '正面', 'negative': '负面', 'neutral': '中性'}

# ── Sheet 1: 全量评论明细 ──
ws1 = wb.active
ws1.title = "评论-店铺-标签明细"

headers = [
    "序号", "店铺名称", "店铺类型", "来源平台", "评分",
    "日期", "评论内容",
    "场景大类", "场景中类", "场景细类", "置信度",
    "情感判定", "LLM备注"
]

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# 写入数据
prev_shop = None
shop_color_idx = 0
shop_colors = ['F5F5F5', 'FFFFFF']  # 交替灰白底色区分商家

for i, r in enumerate(data, 2):
    shop = r.get('shop_name', '')
    if shop != prev_shop:
        shop_color_idx = 1 - shop_color_idx
        prev_shop = shop

    row_fill = PatternFill(start_color=shop_colors[shop_color_idx],
                           end_color=shop_colors[shop_color_idx], fill_type="solid")

    sent = r.get('merged_sentiment', 'neutral')
    rating = r.get('rating', '')
    if rating and isinstance(rating, (int, float)):
        # 携程10分制转5分
        if rating > 5:
            rating = round(rating / 2, 1)

    row_data = [
        i - 1,
        shop,
        r.get('business_type', ''),
        r.get('platform', ''),
        rating,
        r.get('date', ''),
        r.get('content', ''),
        r.get('llm_primary', ''),
        r.get('llm_secondary', ''),
        r.get('llm_tertiary', ''),
        r.get('llm_confidence', ''),
        sent_labels.get(sent, sent),
        r.get('llm_note', ''),
    ]

    for col, val in enumerate(row_data, 1):
        cell = ws1.cell(row=i, column=col, value=val)
        cell.border = thin_border
        cell.font = body_font
        cell.fill = row_fill

        if col == 7:  # 评论内容列
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        elif col == 12:  # 情感列
            cell.fill = sent_fills.get(sent, row_fill)
            cell.alignment = Alignment(horizontal='center')
        elif col in (1, 5, 11):
            cell.alignment = Alignment(horizontal='center')

# 列宽
col_widths = [5, 22, 8, 12, 6, 12, 55, 12, 14, 20, 7, 8, 30]
for col, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(col)].width = w

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:M{len(data)+1}"

# ── Sheet 2: 店铺维度汇总 ──
ws2 = wb.create_sheet("店铺维度汇总")

shop_stats = defaultdict(lambda: {
    'count': 0, 'ratings': [], 'positive': 0, 'negative': 0, 'neutral': 0,
    'labels': Counter(), 'platforms': set(), 'type': ''
})

for r in data:
    shop = r['shop_name']
    s = shop_stats[shop]
    s['count'] += 1
    s['type'] = r.get('business_type', '')
    s['platforms'].add(r.get('platform', ''))

    rating = r.get('rating', '')
    if rating and isinstance(rating, (int, float)):
        v = rating / 2 if rating > 5 else rating
        s['ratings'].append(v)

    sent = r.get('merged_sentiment', 'neutral')
    s[sent] += 1

    for lbl in [r.get('llm_primary', ''), r.get('llm_secondary', ''), r.get('llm_tertiary', '')]:
        if lbl:
            s['labels'][lbl] += 1

# 表头
h2 = ["店铺名称", "类型", "评论数", "均分(5分制)", "正面数", "负面数", "中性数",
      "好评率", "主要场景标签(TOP3)", "来源平台数"]
for col, h in enumerate(h2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 按评论数降序
row = 2
for shop, s in sorted(shop_stats.items(), key=lambda x: -x[1]['count']):
    avg_rating = round(sum(s['ratings']) / len(s['ratings']), 2) if s['ratings'] else ''
    pos_rate = f"{s['positive'] / s['count'] * 100:.0f}%" if s['count'] > 0 else ''
    top3 = ', '.join(f"{k}({v})" for k, v in s['labels'].most_common(3))

    row_data = [
        shop, s['type'], s['count'], avg_rating,
        s['positive'], s['negative'], s['neutral'],
        pos_rate, top3, len(s['platforms'])
    ]

    for col, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.font = body_font
        if col == 9:
            cell.alignment = Alignment(wrap_text=True)
    row += 1

col_widths2 = [22, 8, 8, 10, 8, 8, 8, 8, 40, 10]
for col, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.freeze_panes = "A2"

# ── Sheet 3: 场景标签矩阵 ──
ws3 = wb.create_sheet("场景标签矩阵")

# 三级标签交叉统计
label_matrix = defaultdict(lambda: defaultdict(int))
for r in data:
    p = r.get('llm_primary', '')
    s = r.get('llm_secondary', '')
    t = r.get('llm_tertiary', '')
    if p and s:
        label_matrix[p][f"{s} > {t}"] += 1

ws3.cell(row=1, column=1, value="场景大类").font = Font(bold=True, size=11, name='Microsoft YaHei')
ws3.cell(row=1, column=2, value="场景中类 > 细类").font = Font(bold=True, size=11, name='Microsoft YaHei')
ws3.cell(row=1, column=3, value="评论数").font = Font(bold=True, size=11, name='Microsoft YaHei')
ws3.cell(row=1, column=4, value="占比").font = Font(bold=True, size=11, name='Microsoft YaHei')

for col in range(1, 5):
    ws3.cell(row=1, column=col).fill = hdr_fill
    ws3.cell(row=1, column=col).font = hdr_font
    ws3.cell(row=1, column=col).border = thin_border

row = 2
for primary in sorted(label_matrix.keys()):
    sub_items = sorted(label_matrix[primary].items(), key=lambda x: -x[1])
    for sub_label, count in sub_items:
        ws3.cell(row=row, column=1, value=primary).border = thin_border
        ws3.cell(row=row, column=1).font = body_font
        ws3.cell(row=row, column=2, value=sub_label).border = thin_border
        ws3.cell(row=row, column=2).font = body_font
        ws3.cell(row=row, column=3, value=count).border = thin_border
        ws3.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws3.cell(row=row, column=4, value=f"{count/len(data)*100:.1f}%").border = thin_border
        ws3.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        row += 1

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 8
ws3.column_dimensions['D'].width = 8

# ── Sheet 4: 同比环比 ──
ws4 = wb.create_sheet("同比环比")

# 按年份+大类
year_cat = defaultdict(lambda: defaultdict(int))
for r in data:
    y = r.get('year')
    p = r.get('llm_primary', '')
    if y and p:
        year_cat[y][p] += 1

ws4.cell(row=1, column=1, value="2025 vs 2026 同比分析").font = Font(bold=True, size=13, name='Microsoft YaHei')

h4 = ["场景大类", "2025年", "2026年", "同比变化", "趋势"]
for col, h in enumerate(h4, 1):
    cell = ws4.cell(row=3, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = thin_border

all_cats = set()
for y in year_cat:
    all_cats.update(year_cat[y].keys())

row = 4
for cat in sorted(all_cats):
    c25 = year_cat.get(2025, {}).get(cat, 0)
    c26 = year_cat.get(2026, {}).get(cat, 0)
    chg = ((c26 - c25) / c25 * 100) if c25 > 0 else (100.0 if c26 > 0 else 0)
    trend = "上升" if chg > 10 else ("下降" if chg < -10 else "持平")

    ws4.cell(row=row, column=1, value=cat).border = thin_border
    ws4.cell(row=row, column=2, value=c25).border = thin_border
    ws4.cell(row=row, column=3, value=c26).border = thin_border
    ws4.cell(row=row, column=4, value=f"{chg:+.1f}%").border = thin_border
    ws4.cell(row=row, column=5, value=trend).border = thin_border

    if chg > 50:
        ws4.cell(row=row, column=5).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    elif chg < -30:
        ws4.cell(row=row, column=5).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    row += 1

# 月度环比
row += 2
ws4.cell(row=row, column=1, value="月度环比分析").font = Font(bold=True, size=13, name='Microsoft YaHei')
row += 1

h_mom = ["月份", "总评论", "正面", "负面", "负面占比", "环比变化"]
for col, h in enumerate(h_mom, 1):
    ws4.cell(row=row, column=col, value=h).font = Font(bold=True, name='Microsoft YaHei')
row += 1

month_data = defaultdict(lambda: {"total": 0, "positive": 0, "negative": 0})
for r in data:
    if r.get('year') and r.get('month'):
        key = f"{r['year']}-{r['month']:02d}"
        month_data[key]["total"] += 1
        sent = r.get('merged_sentiment', 'neutral')
        if sent in ('positive', 'negative'):
            month_data[key][sent] += 1

prev_neg = None
for mk in sorted(month_data.keys()):
    d = month_data[mk]
    neg_pct = f"{d['negative']/d['total']*100:.1f}%" if d['total'] > 0 else ''
    chg_str = ""
    if prev_neg is not None:
        chg = d['negative'] - prev_neg
        chg_str = f"{chg:+d}"
    ws4.cell(row=row, column=1, value=mk)
    ws4.cell(row=row, column=2, value=d['total'])
    ws4.cell(row=row, column=3, value=d['positive'])
    ws4.cell(row=row, column=4, value=d['negative'])
    ws4.cell(row=row, column=5, value=neg_pct)
    ws4.cell(row=row, column=6, value=chg_str)
    prev_neg = d['negative']
    row += 1

for col in range(1, 7):
    ws4.column_dimensions[get_column_letter(col)].width = 14

# ── 保存 ──
OUT = r'C:\Users\Administrator\.qoderwork\workspace\mritsdpuqphp34kx\outputs'
excel_path = os.path.join(OUT, '桂山岛评论分析.xlsx')
wb.save(excel_path)
print(f"Excel saved: {excel_path}")

# 同步到项目目录
import shutil
shutil.copy2(excel_path, r'E:\珠海桂山岛案例\数据采集\processed\桂山岛评论分析.xlsx')
print("同步到项目目录完成")

# 汇总
print(f"\n{'='*50}")
print(f"4个Sheet:")
print(f"  1. 评论-店铺-标签明细: {len(data)}条 (按商家分组+交替底色)")
print(f"  2. 店铺维度汇总: {len(shop_stats)}家店铺 (含TOP3标签)")
print(f"  3. 场景标签矩阵: 三级标签交叉统计")
print(f"  4. 同比环比: 2025vs2026 + 月度趋势")
