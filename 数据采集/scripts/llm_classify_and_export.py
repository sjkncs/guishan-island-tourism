# -*- coding: utf-8 -*-
"""
桂山岛评论 LLM 三级场景标签分类 + Excel 导出
==============================================
1. 用 LLM (GPT-5.5) 对每条评论做大/中/细三级分类
2. 合并关键词标签 + LLM标签
3. 导出到 Excel（含同比环比分析）
"""
import json, os, sys, time, hashlib
from datetime import datetime
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

# ── 配置 ──
BASE_DIR = r"E:\珠海桂山岛案例\数据采集"
IN_FILE = os.path.join(BASE_DIR, "processed", "all_reviews.json")
OUT_DIR = os.path.join(BASE_DIR, "processed")

# LLM API 配置 (OpenAI-compatible)
LLM_CONFIGS = [
    {
        "name": "lanyiapi",
        "base_url": "https://lanyiapi.com/v1",
        "api_key": "sk-vpBVC7bc3t9dDAVMPprGsxW4fEgQtuzn2lkMgrGW7SpQGRel",
        "model": "gpt-5.5",
    },
]

# ── 三级分类体系 ──
TAXONOMY = """
请对以下用户评论进行场景标签分类。分类体系如下：

一级（大类）    二级（中类）              三级（细类）
─────────────────────────────────────────────────────
食品安全        食品质量                不新鲜/变质/异味
                                        海鲜质量差(死海鲜/注水/缺斤少两)
                                        以次充好(养殖冒充野生)
              食品卫生                异物(头发/虫子/塑料等)
                                        操作间不卫生/餐具未消毒
                                        生熟混放/过期食品
              食品安全事件            食物中毒/肠胃不适
                                        过敏反应

店铺环境        卫生状况                脏乱(垃圾/油污/灰尘/发霉)
                                        卫生间脏/异味
                                        床单被套不洁(毛发/污渍)
              虫害问题                蟑螂(活体/尸体)
                                        蚊虫/苍蝇/蚂蚁
                                        老鼠
              设施设备                设施陈旧(破损/掉漆/生锈)
                                        空调故障(不制冷/噪音)
                                        热水问题(发黄/不热/水压低)
                                        隔音差(听到隔壁/走廊噪音)
              空间舒适度            空间狭小(放不下行李/拥挤)
                                        潮湿/通风差

服务态度        服务行为                态度恶劣(凶/甩脸/不耐烦)
                                        服务不及时(等很久/叫不应)
                                        服务失误(上错菜/忘单)
              诚信问题                虚假宣传(照骗/到店无房/货不对板)
                                        隐形消费(未标价/结账加价)
                                        退改纠纷(不退钱/霸王条款)

性价比          价格合理性            价格虚高/宰客
                                        节假日涨价(周末/假期翻倍)
                                        性价比低(贵但不值)
              计量诚信                缺斤少两/偷换
                                        分量不足

体验内容        游玩体验                内容匮乏(没啥好玩/无聊)
                                        商业化过重(失去本味)
                                        文化体验缺失
              自然环境                风景优美(海美/日出日落)
                                        海水水质差(不干净/有垃圾)
                                        天气影响(台风/雨天无替代)

交通出行        岛际交通                船班少/等船久
                                        晕船/风浪大
                                        航线不便
              岛上交通                无公交/步行累
                                        道路状况差

住宿体验        住宿品质                超出预期(惊喜/升级/贴心)
                                        符合预期(正常/还行)
                                        低于预期(失望/与宣传不符)

正面体验        推荐意愿                强烈推荐/会再来
                                        一般推荐/看情况
                                        不推荐/踩坑避雷
              美食体验                海鲜好吃/新鲜
                                        特色美食推荐
                                        餐饮一般/不推荐

情感表达        积极情感                开心/满足/惊喜/感动
              消极情感                失望/愤怒/后悔/恶心
              中性表达                客观描述/信息分享

输出格式要求（严格JSON）：
{"primary": "一级大类", "secondary": "二级中类", "tertiary": "三级细类", "confidence": 0.0-1.0, "sentiment": "positive/negative/neutral", "note": "简要说明(可选)"}

如果评论涉及多个场景，选择最主要的一个。confidence表示你对分类的确信度。
"""

def test_api(config):
    """测试API连通性"""
    import urllib.request
    url = f"{config['base_url']}/models"
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {config['api_key']}",
        "Content-Type": "application/json"
    })
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
            models = [m.get('id', '') for m in data.get('data', [])]
            gpt_models = [m for m in models if 'gpt' in m.lower() or '5.5' in m]
            print(f"  [{config['name']}] OK - {len(models)} models, GPT: {gpt_models[:5]}")
            return True, config, gpt_models
    except Exception as e:
        print(f"  [{config['name']}] FAIL - {e}")
        return False, config, []

def call_llm(config, review_content, max_retries=3):
    """调用LLM进行分类"""
    import urllib.request
    
    prompt = f"""{TAXONOMY}

用户评论：
「{review_content}」

请输出JSON分类结果："""

    payload = json.dumps({
        "model": config["model"],
        "messages": [
            {"role": "system", "content": "你是一个专业的旅游评论分析助手，擅长对用户评论进行精准的场景标签分类。严格按JSON格式输出。"},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.1,
        "max_tokens": 200,
    })
    
    for attempt in range(max_retries):
        try:
            req = urllib.request.Request(
                f"{config['base_url']}/chat/completions",
                data=payload.encode('utf-8'),
                headers={
                    "Authorization": f"Bearer {config['api_key']}",
                    "Content-Type": "application/json"
                },
                method="POST"
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                result = json.loads(resp.read())
                content = result['choices'][0]['message']['content']
                # 提取JSON
                content = content.strip()
                if content.startswith('```'):
                    content = content.split('\n', 1)[1].rsplit('```', 1)[0].strip()
                if content.startswith('{'):
                    return json.loads(content)
                # 尝试找到JSON块
                start = content.find('{')
                end = content.rfind('}') + 1
                if start >= 0 and end > start:
                    return json.loads(content[start:end])
                return {"error": "no_json", "raw": content}
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(1)
            else:
                return {"error": str(e)}
    return {"error": "max_retries"}

def batch_classify(reviews, config, batch_size=5, max_workers=3):
    """批量分类（带速率控制）"""
    results = []
    total = len(reviews)
    
    for i in range(0, total, batch_size):
        batch = reviews[i:i+batch_size]
        batch_results = {}
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {}
            for j, review in enumerate(batch):
                idx = i + j
                future = executor.submit(call_llm, config, review['content'])
                futures[future] = idx
            
            for future in as_completed(futures):
                idx = futures[future]
                result = future.result()
                batch_results[idx] = result
        
        results.extend([(idx, batch_results[idx]) for idx in sorted(batch_results)])
        
        done = min(i + batch_size, total)
        errors = sum(1 for _, r in batch_results.items() if 'error' in r)
        sys.stdout.write(f"\r  进度: {done}/{total} ({done/total*100:.1f}%) | 错误: {errors}")
        sys.stdout.flush()
        
        # 速率控制
        if done < total:
            time.sleep(0.5)
    
    print()
    return dict(results)

def merge_labels(reviews, keyword_labels, llm_labels):
    """合并关键词标签和LLM标签"""
    for i, review in enumerate(reviews):
        # 关键词标签
        kw = keyword_labels.get(i, {})
        review['keyword_labels'] = kw.get('labels', [])
        review['keyword_details'] = kw.get('details', {})
        
        # LLM标签
        llm = llm_labels.get(i, {})
        review['llm_primary'] = llm.get('primary', '')
        review['llm_secondary'] = llm.get('secondary', '')
        review['llm_tertiary'] = llm.get('tertiary', '')
        review['llm_confidence'] = llm.get('confidence', 0)
        review['llm_sentiment'] = llm.get('sentiment', '')
        review['llm_note'] = llm.get('note', '')
        
        # 合并情感
        if review['llm_sentiment']:
            review['merged_sentiment'] = review['llm_sentiment']
        else:
            review['merged_sentiment'] = review.get('sentiment', 'neutral')
    
    return reviews

def export_to_excel(reviews, filepath):
    """导出到Excel"""
    try:
        import openpyxl
    except ImportError:
        os.system(f'{sys.executable} -m pip install openpyxl')
        import openpyxl
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # ── Sheet 1: 全量数据 ──
    ws1 = wb.active
    ws1.title = "全量评论数据"
    
    headers = [
        "序号", "平台", "商家名称", "商家类型", "评分",
        "日期", "年份", "月份", "评论内容",
        "关键词-大类", "关键词-中类", "关键词-细类",
        "LLM-大类", "LLM-中类", "LLM-细类", "LLM置信度",
        "合并情感", "LLM情感", "LLM备注"
    ]
    
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 情感颜色
    sentiment_fills = {
        'positive': PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        'negative': PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
        'neutral': PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
    }
    
    for i, r in enumerate(reviews, 2):
        row_data = [
            i - 1,
            r.get('platform', ''),
            r.get('business_name', ''),
            r.get('business_type', ''),
            r.get('rating', ''),
            r.get('date', ''),
            r.get('year', ''),
            r.get('month', ''),
            r.get('content', '')[:500],  # 限制长度
            # 关键词标签（如果有）
            r.get('keyword_labels', [''])[0] if r.get('keyword_labels') else '',
            '',  # keyword secondary (from keyword_details)
            '',  # keyword tertiary
            r.get('llm_primary', ''),
            r.get('llm_secondary', ''),
            r.get('llm_tertiary', ''),
            r.get('llm_confidence', ''),
            r.get('merged_sentiment', ''),
            r.get('llm_sentiment', ''),
            r.get('llm_note', ''),
        ]
        
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col == 9))
            
            # 情感行底色
            sent = r.get('merged_sentiment', 'neutral')
            if sent in sentiment_fills:
                if col == 17:  # 合并情感列
                    cell.fill = sentiment_fills[sent]
    
    # 列宽
    col_widths = [5, 12, 18, 8, 6, 12, 6, 6, 50, 12, 12, 12, 12, 14, 18, 8, 10, 10, 20]
    for col, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
    
    # 冻结首行
    ws1.freeze_panes = "A2"
    
    # ── Sheet 2: 标签统计 ──
    ws2 = wb.create_sheet("标签统计")
    
    # LLM标签分布
    ws2.cell(row=1, column=1, value="LLM三级标签分布统计").font = Font(bold=True, size=14)
    
    label_counts = defaultdict(lambda: defaultdict(int))
    for r in reviews:
        p = r.get('llm_primary', '')
        s = r.get('llm_secondary', '')
        t = r.get('llm_tertiary', '')
        if p:
            label_counts[p][f"{s} > {t}"] += 1
    
    row = 3
    ws2.cell(row=row, column=1, value="一级大类").font = Font(bold=True)
    ws2.cell(row=row, column=2, value="二级中类 > 三级细类").font = Font(bold=True)
    ws2.cell(row=row, column=3, value="数量").font = Font(bold=True)
    ws2.cell(row=row, column=4, value="占比").font = Font(bold=True)
    row += 1
    
    total_reviews = len(reviews)
    for primary in sorted(label_counts.keys()):
        for secondary_tertiary, count in sorted(label_counts[primary].items(), key=lambda x: -x[1]):
            ws2.cell(row=row, column=1, value=primary)
            ws2.cell(row=row, column=2, value=secondary_tertiary)
            ws2.cell(row=row, column=3, value=count)
            ws2.cell(row=row, column=4, value=f"{count/total_reviews*100:.1f}%")
            row += 1
    
    # 列宽
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 8
    ws2.column_dimensions['D'].width = 8
    
    # ── Sheet 3: 同比环比 ──
    ws3 = wb.create_sheet("同比环比分析")
    
    ws3.cell(row=1, column=1, value="2025 vs 2026 同比分析").font = Font(bold=True, size=14)
    
    # 按年份+大类统计
    year_labels = defaultdict(lambda: defaultdict(int))
    for r in reviews:
        y = r.get('year')
        p = r.get('llm_primary', '')
        if y and p:
            year_labels[y][p] += 1
    
    row = 3
    headers3 = ["一级大类", "2025年", "2026年", "同比变化", "趋势"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    row += 1
    
    all_categories = set()
    for y in year_labels:
        all_categories.update(year_labels[y].keys())
    
    for cat in sorted(all_categories):
        c25 = year_labels.get(2025, {}).get(cat, 0)
        c26 = year_labels.get(2026, {}).get(cat, 0)
        change = ((c26 - c25) / c25 * 100) if c25 > 0 else (100.0 if c26 > 0 else 0)
        trend = "上升" if change > 10 else ("下降" if change < -10 else "持平")
        
        ws3.cell(row=row, column=1, value=cat)
        ws3.cell(row=row, column=2, value=c25)
        ws3.cell(row=row, column=3, value=c26)
        ws3.cell(row=row, column=4, value=f"{change:+.1f}%")
        ws3.cell(row=row, column=5, value=trend)
        
        if change > 50:
            ws3.cell(row=row, column=5).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        elif change < -30:
            ws3.cell(row=row, column=5).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        row += 1
    
    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 10
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 8
    
    # 环比
    row += 2
    ws3.cell(row=row, column=1, value="月度环比分析").font = Font(bold=True, size=14)
    row += 1
    
    month_data = defaultdict(lambda: {"total": 0, "negative": 0, "positive": 0})
    for r in reviews:
        if r.get('year') and r.get('month'):
            key = f"{r['year']}-{r['month']:02d}"
            month_data[key]["total"] += 1
            sent = r.get('merged_sentiment', 'neutral')
            if sent == 'negative':
                month_data[key]["negative"] += 1
            elif sent == 'positive':
                month_data[key]["positive"] += 1
    
    headers_mom = ["月份", "总评论", "正面", "负面", "负面占比", "环比变化"]
    for col, h in enumerate(headers_mom, 1):
        cell = ws3.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
    row += 1
    
    prev_neg = None
    for month_key in sorted(month_data.keys()):
        d = month_data[month_key]
        neg_pct = d["negative"] / d["total"] * 100 if d["total"] > 0 else 0
        
        change_str = ""
        if prev_neg is not None:
            change = d["negative"] - prev_neg
            change_str = f"{change:+d}"
        
        ws3.cell(row=row, column=1, value=month_key)
        ws3.cell(row=row, column=2, value=d["total"])
        ws3.cell(row=row, column=3, value=d["positive"])
        ws3.cell(row=row, column=4, value=d["negative"])
        ws3.cell(row=row, column=5, value=f"{neg_pct:.1f}%")
        ws3.cell(row=row, column=6, value=change_str)
        
        prev_neg = d["negative"]
        row += 1
    
    # ── Sheet 4: 长尾高危 ──
    ws4 = wb.create_sheet("长尾高危标签")
    
    ws4.cell(row=1, column=1, value="长尾高危标签（占比<5%但严重程度高）").font = Font(bold=True, size=14)
    
    high_severity_keywords = {
        '食品': ['不新鲜', '变质', '中毒', '异物', '蟑螂', '苍蝇', '过期'],
        '安全': ['中毒', '过敏', '呕吐', '不适', '受伤', '事故'],
        '诚信': ['宰客', '霸王', '不退', '欺诈', '偷换', '缺斤少两'],
    }
    
    row = 3
    headers_lt = ["评论内容", "LLM大类", "LLM中类", "LLM细类", "情感", "高危原因"]
    for col, h in enumerate(headers_lt, 1):
        cell = ws4.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    row += 1
    
    for r in reviews:
        content = r.get('content', '')
        is_high_risk = False
        risk_reasons = []
        
        for category, keywords in high_severity_keywords.items():
            for kw in keywords:
                if kw in content:
                    is_high_risk = True
                    risk_reasons.append(f"{category}:{kw}")
        
        if is_high_risk and r.get('merged_sentiment') == 'negative':
            ws4.cell(row=row, column=1, value=content[:200])
            ws4.cell(row=row, column=2, value=r.get('llm_primary', ''))
            ws4.cell(row=row, column=3, value=r.get('llm_secondary', ''))
            ws4.cell(row=row, column=4, value=r.get('llm_tertiary', ''))
            ws4.cell(row=row, column=5, value=r.get('merged_sentiment', ''))
            ws4.cell(row=row, column=6, value=', '.join(risk_reasons))
            
            for col in range(1, 7):
                ws4.cell(row=row, column=col).fill = PatternFill(
                    start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
                )
            row += 1
    
    ws4.column_dimensions['A'].width = 50
    ws4.column_dimensions['B'].width = 12
    ws4.column_dimensions['C'].width = 14
    ws4.column_dimensions['D'].width = 18
    ws4.column_dimensions['E'].width = 10
    ws4.column_dimensions['F'].width = 25
    
    # 保存
    wb.save(filepath)
    print(f"[EXPORT] Excel saved to: {filepath}")
    return filepath


def main():
    sys.stdout.reconfigure(encoding='utf-8')
    print("=" * 60)
    print(" 桂山岛 LLM 三级场景标签分类 + Excel 导出")
    print("=" * 60)
    
    # 1. 加载数据
    print("\n[1/5] 加载评论数据...")
    with open(IN_FILE, 'r', encoding='utf-8') as f:
        reviews = json.load(f)
    print(f"  {len(reviews)} 条评论")
    
    # 2. 测试API
    print("\n[2/5] 测试 LLM API 连通性...")
    working_config = None
    
    for config in LLM_CONFIGS:
        if not config['api_key']:
            print(f"  [{config['name']}] SKIP - 未配置 API key")
            continue
        ok, _, models = test_api(config)
        if ok:
            working_config = config
            # 尝试找到最佳模型
            for m in models:
                if '5.5' in m or 'gpt-5' in m:
                    config['model'] = m
                    break
            break
    
    if not working_config:
        print("\n  未找到可用的 LLM API。")
        print("  请设置环境变量后重试：")
        print("    set MYDAMOXING_API_KEY=your-key-here")
        print("    set XLAPIS_API_KEY=your-key-here")
        print("\n  或手动编辑脚本中的 LLM_CONFIGS 填入 API key。")
        
        # 即使没有LLM，也导出Excel（仅含关键词标签）
        print("\n[4/5] 跳过LLM分类，直接导出Excel（仅关键词标签）...")
        excel_path = os.path.join(OUT_DIR, "桂山岛评论分析.xlsx")
        
        # 用关键词标签填充LLM字段
        from guishan_collector import ScenarioLabeler
        labeler = ScenarioLabeler()
        for r in reviews:
            labels, details = labeler.label_review(r.get('content', ''))
            r['keyword_labels'] = labels
            r['keyword_details'] = details
            # 用关键词标签作为LLM标签的fallback
            if labels:
                first_label = labels[0]
                info = labeler.labels.get(first_label, {})
                r['llm_primary'] = info.get('category', '')
                r['llm_secondary'] = info.get('label', '')
                r['llm_tertiary'] = ''
            else:
                r['llm_primary'] = ''
                r['llm_secondary'] = ''
                r['llm_tertiary'] = ''
            r['llm_confidence'] = 0
            r['llm_sentiment'] = ''
            r['llm_note'] = '关键词匹配fallback'
            
            # 情感
            neg = sum(1 for l in labels if not l.startswith('G'))
            pos = sum(1 for l in labels if l.startswith('G'))
            if neg > pos:
                r['merged_sentiment'] = 'negative'
            elif pos > neg:
                r['merged_sentiment'] = 'positive'
            else:
                r['merged_sentiment'] = r.get('sentiment', 'neutral')
        
        export_to_excel(reviews, excel_path)
        return
    
    # 3. 批量分类
    print(f"\n[3/5] LLM 批量分类 ({len(reviews)}条评论)...")
    print(f"  使用: {working_config['name']} ({working_config['model']})")
    
    llm_results = batch_classify(reviews, working_config)
    
    # 4. 合并标签
    print("\n[4/5] 合并关键词标签 + LLM标签...")
    from guishan_collector import ScenarioLabeler
    labeler = ScenarioLabeler()
    
    keyword_labels = {}
    for i, r in enumerate(reviews):
        labels, details = labeler.label_review(r.get('content', ''))
        keyword_labels[i] = {'labels': labels, 'details': details}
    
    reviews = merge_labels(reviews, keyword_labels, llm_results)
    
    # 统计LLM分类结果
    llm_ok = sum(1 for i, r in llm_results.items() if 'error' not in r)
    llm_err = sum(1 for i, r in llm_results.items() if 'error' in r)
    print(f"  LLM成功: {llm_ok}/{len(reviews)} | 失败: {llm_err}")
    
    # 5. 导出Excel
    print("\n[5/5] 导出 Excel...")
    excel_path = os.path.join(OUT_DIR, "桂山岛评论分析.xlsx")
    export_to_excel(reviews, excel_path)
    
    # 同时保存增强版JSON
    with open(os.path.join(OUT_DIR, "all_reviews_llm.json"), 'w', encoding='utf-8') as f:
        json.dump(reviews, f, ensure_ascii=False, indent=2)
    print(f"  [EXPORT] all_reviews_llm.json")
    
    print(f"\n{'='*60}")
    print(f" 完成！Excel: {excel_path}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
