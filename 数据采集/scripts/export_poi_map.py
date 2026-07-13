# -*- coding: utf-8 -*-
"""
桂山岛POI数据导出Excel + 叠加底图
"""
import json, os, sys
sys.stdout.reconfigure(encoding='utf-8')

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.offsetbox import OffsetImage, AnnotationBbox
import numpy as np

# ── 中文字体配置 ──
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

BASE_DIR = r'E:\珠海桂山岛案例\数据采集'
POI_FILE = os.path.join(BASE_DIR, 'raw', 'amap_poi_guishan.json')
OUT_DIR = os.path.join(BASE_DIR, 'processed')

# ── 加载数据 ──
with open(POI_FILE, 'r', encoding='utf-8') as f:
    data = json.load(f)
pois = data['pois']
print(f"加载 {len(pois)} 个POI")

# ══════════════════════════════════════════
# Part 1: 导出Excel
# ══════════════════════════════════════════
try:
    import openpyxl
except ImportError:
    os.system(f'{sys.executable} -m pip install openpyxl')
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Sheet 1: 全量POI ──
ws1 = wb.active
ws1.title = "桂山岛POI全量数据"

headers = [
    "序号", "名称", "类别", "细类", "评分", "地址",
    "经度", "纬度", "电话", "商圈", "搜索关键词", "POI ID"
]

# 表头样式
hdr_font = Font(bold=True, color="FFFFFF", size=11, name='Microsoft YaHei')
hdr_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 类别颜色
cat_colors = {
    '住宿': 'E3F2FD',  # 浅蓝
    '景点': 'E8F5E9',  # 浅绿
    '餐饮': 'FFF3E0',  # 浅橙
    '购物': 'F3E5F5',  # 浅紫
    '交通': 'E0F7FA',  # 浅青
    '公共设施': 'FFFDE7', # 浅黄
    '休闲': 'FCE4EC',  # 浅粉
    '生活服务': 'F1F8E9', # 浅草绿
    '汽车服务': 'ECEFF1',
    '政府': 'EFEBE9',
    '教育': 'E8EAF6',
    '地名': 'F5F5F5',
    '医疗': 'FFEBEE',
}

for i, p in enumerate(pois, 2):
    loc = p.get('location', {})
    row_data = [
        i - 1,
        p.get('name', ''),
        p.get('category', ''),
        p.get('type', ''),
        p.get('rating', ''),
        p.get('address', ''),
        loc.get('lng', ''),
        loc.get('lat', ''),
        p.get('tel', ''),
        p.get('business_area', ''),
        p.get('search_keyword', ''),
        p.get('poi_id', ''),
    ]
    
    cat = p.get('category', '')
    fill_color = cat_colors.get(cat, 'FFFFFF')
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    for col, val in enumerate(row_data, 1):
        cell = ws1.cell(row=i, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=(col in [4, 6]))
        cell.fill = row_fill
        cell.font = Font(name='Microsoft YaHei', size=10)

# 列宽
col_widths = [5, 28, 8, 30, 6, 35, 12, 12, 15, 10, 18, 14]
for col, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(col)].width = w

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:L{len(pois)+1}"

# ── Sheet 2: 分类统计 ──
ws2 = wb.create_sheet("分类统计")
ws2.cell(row=1, column=1, value="类别").font = Font(bold=True, name='Microsoft YaHei')
ws2.cell(row=1, column=2, value="数量").font = Font(bold=True, name='Microsoft YaHei')
ws2.cell(row=1, column=3, value="占比").font = Font(bold=True, name='Microsoft YaHei')
ws2.cell(row=1, column=4, value="平均评分").font = Font(bold=True, name='Microsoft YaHei')

from collections import Counter, defaultdict
cat_counts = Counter(p.get('category', '未知') for p in pois)
cat_ratings = defaultdict(list)
for p in pois:
    r = p.get('rating', '')
    if r:
        try:
            cat_ratings[p.get('category', '未知')].append(float(r))
        except ValueError:
            pass

row = 2
for cat, count in cat_counts.most_common():
    ws2.cell(row=row, column=1, value=cat)
    ws2.cell(row=row, column=2, value=count)
    ws2.cell(row=row, column=3, value=f"{count/len(pois)*100:.1f}%")
    ratings = cat_ratings.get(cat, [])
    avg = f"{sum(ratings)/len(ratings):.2f}" if ratings else "N/A"
    ws2.cell(row=row, column=4, value=avg)
    
    fill_color = cat_colors.get(cat, 'FFFFFF')
    for c in range(1, 5):
        ws2.cell(row=row, column=c).fill = PatternFill(
            start_color=fill_color, end_color=fill_color, fill_type="solid"
        )
    row += 1

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 8
ws2.column_dimensions['C'].width = 8
ws2.column_dimensions['D'].width = 10

# 保存Excel
excel_path = os.path.join(OUT_DIR, '桂山岛POI数据.xlsx')
wb.save(excel_path)
print(f"[EXPORT] Excel: {excel_path}")

# ══════════════════════════════════════════
# Part 2: 生成底图叠加图
# ══════════════════════════════════════════

# 过滤桂山岛本体POI (lat 22.12-22.16, lng 113.81-113.85)
island_pois = [p for p in pois 
               if 'location' in p 
               and 22.12 <= p['location']['lat'] <= 22.16
               and 113.81 <= p['location']['lng'] <= 113.85]
other_pois = [p for p in pois if p not in island_pois and 'location' in p]

print(f"岛内POI: {len(island_pois)}, 岛外POI: {len(other_pois)}")

# 类别配色方案
cat_style = {
    '住宿': {'color': '#2196F3', 'marker': 'o', 'size': 60, 'label': '住宿(民宿/酒店)'},
    '景点': {'color': '#4CAF50', 'marker': '*', 'size': 120, 'label': '景点(景区/公园)'},
    '餐饮': {'color': '#FF9800', 'marker': 's', 'size': 70, 'label': '餐饮(餐厅/海鲜)'},
    '购物': {'color': '#9C27B0', 'marker': 'D', 'size': 50, 'label': '购物(商店/超市)'},
    '交通': {'color': '#00BCD4', 'marker': '^', 'size': 80, 'label': '交通(码头/车站)'},
    '公共设施': {'color': '#FFC107', 'marker': 'P', 'size': 60, 'label': '公共设施(厕所/银行)'},
    '休闲': {'color': '#E91E63', 'marker': 'v', 'size': 50, 'label': '休闲(咖啡/娱乐)'},
    '生活服务': {'color': '#8BC34A', 'marker': 'p', 'size': 50, 'label': '生活服务'},
    '医疗': {'color': '#F44336', 'marker': '+', 'size': 80, 'label': '医疗(卫生院)'},
    '教育': {'color': '#3F51B5', 'marker': 'h', 'size': 60, 'label': '教育(学校)'},
    '政府': {'color': '#795548', 'marker': 'X', 'size': 70, 'label': '政府(村委会)'},
    '地名': {'color': '#607D8B', 'marker': '.', 'size': 40, 'label': '地名'},
    '汽车服务': {'color': '#9E9E9E', 'marker': 'd', 'size': 40, 'label': '汽车服务'},
}

# ── 图1: 全岛总览 ──
fig1, ax1 = plt.subplots(1, 1, figsize=(16, 12))
ax1.set_facecolor('#E8F4FD')

# 画岛外POI（灰色小点）
for p in other_pois:
    loc = p['location']
    ax1.scatter(loc['lng'], loc['lat'], c='#BDBDBD', s=20, marker='.', alpha=0.5, zorder=1)

# 画岛内POI
plotted_cats = set()
for p in island_pois:
    cat = p.get('category', '未知')
    style = cat_style.get(cat, {'color': '#757575', 'marker': '.', 'size': 40, 'label': cat})
    loc = p['location']
    
    label = style['label'] if cat not in plotted_cats else None
    plotted_cats.add(cat)
    
    ax1.scatter(loc['lng'], loc['lat'], 
                c=style['color'], s=style['size'], marker=style['marker'],
                alpha=0.85, zorder=3, label=label, edgecolors='white', linewidth=0.5)

# 标注重要POI名称
key_pois = ['桂山岛风景区', '桂山灯塔', '桂山舰纪念公园', '万山海战遗址', 
            '妈祖庙', '文天祥广场', '一湾沙滩', '桂山码头', '桂山客运站',
            '桂山小学', '桂山镇中心卫生院', '桂山岛']
for p in island_pois:
    name = p.get('name', '')
    if any(k in name for k in key_pois):
        loc = p['location']
        ax1.annotate(name, (loc['lng'], loc['lat']),
                    fontsize=8, fontweight='bold', color='#1A2744',
                    xytext=(5, 5), textcoords='offset points',
                    bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.8, edgecolor='#CCCCCC'),
                    zorder=5)

# 桂山岛区域框
ax1.axhspan(22.12, 22.16, xmin=0.1, xmax=0.9, alpha=0.08, color='#4CAF50', zorder=0)
ax1.axvspan(113.81, 113.85, ymin=0.1, ymax=0.9, alpha=0.08, color='#4CAF50', zorder=0)

ax1.set_xlabel('经度 (Longitude)', fontsize=12, fontname='Microsoft YaHei')
ax1.set_ylabel('纬度 (Latitude)', fontsize=12, fontname='Microsoft YaHei')
ax1.set_title('珠海桂山岛 POI 空间分布总览\n(115个兴趣点 · 高德地图API采集)', 
              fontsize=16, fontweight='bold', fontname='Microsoft YaHei', pad=15)

ax1.legend(loc='lower right', fontsize=9, framealpha=0.9, ncol=2,
          title='POI类别', title_fontsize=10)
ax1.grid(True, alpha=0.3, linestyle='--')

# 比例尺标注
ax1.annotate('≈1km', xy=(113.84, 22.125), fontsize=9, color='#666666',
            ha='center', fontname='Microsoft YaHei')
ax1.plot([113.835, 113.845], [22.126, 22.126], 'k-', linewidth=2)

plt.tight_layout()
fig1_path = os.path.join(OUT_DIR, '桂山岛POI分布_全岛总览.png')
fig1.savefig(fig1_path, dpi=200, bbox_inches='tight', facecolor='white')
plt.close(fig1)
print(f"[EXPORT] 全岛总览图: {fig1_path}")

# ── 图2: 岛内放大详图 ──
fig2, ax2 = plt.subplots(1, 1, figsize=(16, 14))
ax2.set_facecolor('#E8F4FD')

# 只画岛内POI
plotted_cats2 = set()
for p in island_pois:
    cat = p.get('category', '未知')
    style = cat_style.get(cat, {'color': '#757575', 'marker': '.', 'size': 40, 'label': cat})
    loc = p['location']
    
    label = style['label'] if cat not in plotted_cats2 else None
    plotted_cats2.add(cat)
    
    ax2.scatter(loc['lng'], loc['lat'], 
                c=style['color'], s=style['size'] * 2, marker=style['marker'],
                alpha=0.85, zorder=3, label=label, edgecolors='white', linewidth=0.8)

# 标注所有岛内POI名称（岛内POI不多，可以全部标注）
for p in island_pois:
    name = p.get('name', '')
    loc = p['location']
    cat = p.get('category', '')
    
    # 缩短名称
    short_name = name.replace('珠海桂山岛', '').replace('桂山岛', '').replace('(珠海桂山岛店)', '')
    if len(short_name) > 12:
        short_name = short_name[:12] + '..'
    
    fontsize = 7 if cat == '住宿' else 8
    
    ax2.annotate(short_name, (loc['lng'], loc['lat']),
                fontsize=fontsize, color='#333333',
                xytext=(3, 3), textcoords='offset points',
                bbox=dict(boxstyle='round,pad=0.2', facecolor='white', alpha=0.7, edgecolor='#DDDDDD'),
                zorder=5)

# 设置范围为桂山岛本体
ax2.set_xlim(113.815, 113.840)
ax2.set_ylim(22.125, 22.155)

ax2.set_xlabel('经度 (Longitude)', fontsize=12, fontname='Microsoft YaHei')
ax2.set_ylabel('纬度 (Latitude)', fontsize=12, fontname='Microsoft YaHei')
ax2.set_title('桂山岛本体 POI 详细分布\n(岛内兴趣点放大视图)', 
              fontsize=16, fontweight='bold', fontname='Microsoft YaHei', pad=15)

ax2.legend(loc='lower right', fontsize=9, framealpha=0.9, ncol=2,
          title='POI类别', title_fontsize=10)
ax2.grid(True, alpha=0.3, linestyle='--')

# 指北针
ax2.annotate('N', xy=(113.837, 22.152), fontsize=14, fontweight='bold', 
            ha='center', color='#1A2744')
ax2.annotate('', xy=(113.837, 22.154), xytext=(113.837, 22.150),
            arrowprops=dict(arrowstyle='->', color='#1A2744', lw=2))

plt.tight_layout()
fig2_path = os.path.join(OUT_DIR, '桂山岛POI分布_岛内详图.png')
fig2.savefig(fig2_path, dpi=200, bbox_inches='tight', facecolor='white')
plt.close(fig2)
print(f"[EXPORT] 岛内详图: {fig2_path}")

# ── 图3: 分类饼图 + 评分分布 ──
fig3, (ax3a, ax3b) = plt.subplots(1, 2, figsize=(16, 7))

# 饼图
cats_sorted = cat_counts.most_common()
labels = [f"{c}\n({n})" for c, n in cats_sorted]
sizes = [n for _, n in cats_sorted]
colors = [cat_style.get(c, {'color': '#757575'})['color'] for c, _ in cats_sorted]

wedges, texts, autotexts = ax3a.pie(sizes, labels=labels, colors=colors, 
                                      autopct='%1.1f%%', startangle=90,
                                      textprops={'fontsize': 9, 'fontname': 'Microsoft YaHei'})
for t in autotexts:
    t.set_fontsize(8)
ax3a.set_title('POI类别分布', fontsize=14, fontweight='bold', fontname='Microsoft YaHei')

# 评分分布
ratings = []
for p in pois:
    r = p.get('rating', '')
    if r:
        try:
            ratings.append(float(r))
        except ValueError:
            pass

if ratings:
    ax3b.hist(ratings, bins=20, color='#2196F3', edgecolor='white', alpha=0.8)
    ax3b.axvline(np.mean(ratings), color='#F44336', linestyle='--', linewidth=2,
                label=f'均值: {np.mean(ratings):.2f}')
    ax3b.axvline(np.median(ratings), color='#4CAF50', linestyle='--', linewidth=2,
                label=f'中位数: {np.median(ratings):.2f}')
    ax3b.set_xlabel('评分', fontsize=12, fontname='Microsoft YaHei')
    ax3b.set_ylabel('POI数量', fontsize=12, fontname='Microsoft YaHei')
    ax3b.set_title('POI评分分布', fontsize=14, fontweight='bold', fontname='Microsoft YaHei')
    ax3b.legend(fontsize=10)

plt.tight_layout()
fig3_path = os.path.join(OUT_DIR, '桂山岛POI_分类与评分分析.png')
fig3.savefig(fig3_path, dpi=200, bbox_inches='tight', facecolor='white')
plt.close(fig3)
print(f"[EXPORT] 分析图: {fig3_path}")

# ── 复制到用户目录 ──
import shutil
outputs_dir = r'C:\Users\Administrator\.qoderwork\workspace\mritsdpuqphp34kx\outputs'
os.makedirs(outputs_dir, exist_ok=True)
for f in ['桂山岛POI数据.xlsx', '桂山岛POI分布_全岛总览.png', '桂山岛POI分布_岛内详图.png', '桂山岛POI_分类与评分分析.png']:
    src = os.path.join(OUT_DIR, f)
    dst = os.path.join(outputs_dir, f)
    if os.path.exists(src):
        shutil.copy2(src, dst)
        print(f"  → 复制到outputs: {f}")

print("\n全部导出完成！")
